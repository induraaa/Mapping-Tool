"""
Wafer Map Viewer  —  Keithley ACS KDF V1.2
Displays any KDF file as an interactive wafer map.
Subsite number = design number; switch designs via the Design selector.

Requirements:  pip install PySide6
Usage:         python wafer_mapper_light.py [file.kdf]
"""

import sys, os, re, math, statistics
from collections import defaultdict

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QFileDialog, QTreeWidget, QTreeWidgetItem, QGroupBox,
    QLineEdit, QFormLayout, QStatusBar, QComboBox,
    QMessageBox, QTabWidget, QTableWidget, QTableWidgetItem,
    QHeaderView, QToolBar, QSizePolicy, QPushButton, QSpinBox, QCheckBox, QProgressBar,
    QScrollArea, QRadioButton, QButtonGroup,
    QStyle, QStyleOptionComboBox, QStyledItemDelegate
)
from PySide6.QtCore import Qt, QRectF, QPointF, Signal, QSize, QRect
from PySide6.QtGui import (
    QPainter, QColor, QBrush, QPen, QFont,
    QLinearGradient, QRadialGradient, QPixmap, QIcon, QAction,
    QPolygonF, QImage, QImageWriter
)

# ─────────────────────────────────────────────────────────────────────────────
#  THEME
# ─────────────────────────────────────────────────────────────────────────────

T = {
    # Slightly warm off-white theme (less stark than pure white)
    "bg_app":        "#f2f0ec",
    "bg_panel":      "#fbfaf7",
    "bg_header":     "#efeae2",
    "bg_row_alt":    "#f6f2ec",
    "border":        "#cdd6e3",
    "border_hi":     "#a0b4cc",
    "accent":        "#1565c0",
    "accent_dim":    "#dce8fb",
    "accent_dark":   "#0d47a1",
    "pass_bg":       "#e3f6ec",
    "pass_fg":       "#1b6b3a",
    "pass_border":   "#5cb87a",
    "neutral_bg":    "#dfe8f3",
    "neutral_fg":    "#2c4a6e",
    "neutral_border":"#90a8c4",
    "warn_bg":       "#fff6d6",
    "warn_fg":       "#9b7800",
    "warn_border":   "#e5c84a",
    "fail_bg":       "#fce8e8",
    "fail_fg":       "#b71c1c",
    "fail_border":   "#e57373",
    "nodata_bg":     "#eeecf5",
    "nodata_fg":     "#7060a0",
    "nodata_border": "#b0a8d0",
    "wafer_bg":      "#eef2f8",
    "wafer_edge":    "#7a9cbd",
    "text_primary":  "#1a2537",
    "text_secondary":"#4a5c72",
    "text_dim":      "#8fa4bc",
    "selected":      "#e65100",
    "hover_border":  "#1565c0",
    "warn":          "#e65100",
}

INVALID_LIMIT = object()
PASS_COLOR = '#0BDA51'
FAIL_COLOR = '#DC143C'
WARN_COLOR = '#FDDA0D'


def _lerp_color(c1: QColor, c2: QColor, t: float) -> QColor:
    t = max(0.0, min(1.0, float(t)))
    r = int(c1.red() + (c2.red() - c1.red()) * t)
    g = int(c1.green() + (c2.green() - c1.green()) * t)
    b = int(c1.blue() + (c2.blue() - c1.blue()) * t)
    return QColor(r, g, b)

SS = """
QMainWindow,QWidget{
    background-color:""" + T['bg_app'] + """;
    color:""" + T['text_primary'] + """;
    font-family:'Segoe UI','Calibri',sans-serif;
    font-size:13px;
}
QGroupBox{
    background-color:""" + T['bg_panel'] + """;
    border:1px solid """ + T['border'] + """;
    border-radius:10px;
    margin-top:22px;
    padding:10px 8px 8px 8px;
    font-weight:bold;
    font-size:11px;
}
QGroupBox::title{
    subcontrol-origin:margin;
    left:12px;
    padding:2px 8px;
    background:""" + T['bg_panel'] + """;
    color:""" + T['accent'] + """;
    font-size:11px;
    font-weight:bold;
    border-radius:4px;
}
QLabel{background:transparent;color:""" + T['text_primary'] + """;font-size:13px;}
QPushButton{
    background-color:""" + T['bg_panel'] + """;
    border:1px solid """ + T['border_hi'] + """;
    border-radius:8px;
    padding:6px 16px;
    color:""" + T['text_primary'] + """;
    font-weight:600;
    font-size:13px;
    min-height:28px;
}
QPushButton:hover{
    background-color:""" + T['accent_dim'] + """;
    border:1px solid """ + T['accent'] + """;
    color:""" + T['accent_dark'] + """;
}
QPushButton:pressed{background-color:""" + T['accent'] + """;color:white;}
QPushButton#primary{
    background-color:""" + T['accent'] + """;
    border:1px solid """ + T['accent_dark'] + """;
    color:white;
    font-weight:bold;
    border-radius:8px;
}
QPushButton#primary:hover{background-color:""" + T['accent_dark'] + """;}
QComboBox{
    background-color:""" + T['bg_panel'] + """;
    border:1px solid """ + T['border'] + """;
    border-radius:8px;
    padding:5px 36px 5px 12px;
    color:""" + T['text_primary'] + """;
    font-size:13px;
    min-height:32px;
}
QComboBox:hover{border:1px solid """ + T['accent'] + """;}
QComboBox:focus{border:1px solid """ + T['accent'] + """;}
QComboBox::drop-down{
    subcontrol-origin:padding;
    subcontrol-position:right center;
    width:32px;
    border:none;
    border-left:1px solid """ + T['border'] + """;
    border-top-right-radius:8px;
    border-bottom-right-radius:8px;
    background:""" + T['bg_header'] + """;
}
QComboBox::down-arrow{
    image:none;
    width:0px;
    height:0px;
}
QComboBox QAbstractItemView{
    background:""" + T['bg_panel'] + """;
    border:none;
    border-radius:0px;
    selection-background-color:""" + T['accent_dim'] + """;
    color:""" + T['text_primary'] + """;
    font-size:13px;
    padding:4px;
    outline:0;
}
/* ComboBox popup: avoid double frame (common on Windows) */
QComboBoxPrivateContainer{
    background:""" + T['bg_panel'] + """;
    border:1px solid """ + T['border_hi'] + """;
    border-radius:0px;
    padding:0px;
    margin:0px;
}
QComboBoxPrivateContainer QFrame{
    border:none;
}
QComboBoxPrivateContainer QListView{
    border:none;
    background:""" + T['bg_panel'] + """;
    outline:0;
}
QComboBoxPrivateContainer QAbstractScrollArea{
    border:none;
}
QComboBoxPrivateContainer QAbstractScrollArea::viewport{
    background:""" + T['bg_panel'] + """;
    border:none;
}
QComboBoxPrivateContainer *{
    border:none;
    outline:0;
}
QLineEdit,QSpinBox{
    background-color:""" + T['bg_panel'] + """;
    border:1px solid """ + T['border'] + """;
    border-radius:8px;
    padding:5px 10px;
    color:""" + T['text_primary'] + """;
    font-size:13px;
    min-height:28px;
}
QLineEdit:focus,QSpinBox:focus{border:1px solid """ + T['accent'] + """;}
QSpinBox::up-button,QSpinBox::down-button{
    width:20px;
    background:""" + T['bg_header'] + """;
    border:none;
    border-radius:4px;
}
QTreeWidget,QTableWidget{
    background-color:""" + T['bg_panel'] + """;
    border:1px solid """ + T['border'] + """;
    border-radius:8px;
    alternate-background-color:""" + T['bg_row_alt'] + """;
    outline:none;
    font-size:13px;
}
QTreeWidget::item,QTableWidget::item{padding:4px 5px;border:none;}
QTreeWidget::item:hover,QTableWidget::item:hover{
    background-color:""" + T['accent_dim'] + """;
    border-radius:4px;
}
QTreeWidget::item:selected,QTableWidget::item:selected{
    background-color:""" + T['accent_dim'] + """;
    color:""" + T['accent_dark'] + """;
    border-radius:4px;
}
QHeaderView::section{
    background-color:""" + T['bg_header'] + """;
    color:""" + T['accent_dark'] + """;
    border:none;
    border-right:1px solid """ + T['border'] + """;
    border-bottom:1px solid """ + T['border'] + """;
    padding:6px 10px;
    font-size:12px;
    font-weight:bold;
}
QHeaderView::section:first{border-top-left-radius:8px;}
QHeaderView::section:last{border-top-right-radius:8px;border-right:none;}
QScrollBar:vertical{
    background:""" + T['bg_app'] + """;
    width:9px;border:none;border-radius:4px;
}
QScrollBar::handle:vertical{
    background:""" + T['border_hi'] + """;
    border-radius:4px;min-height:24px;
}
QScrollBar::handle:vertical:hover{background:""" + T['accent'] + """;}
QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{height:0;}
QScrollBar:horizontal{
    background:""" + T['bg_app'] + """;
    height:9px;border-radius:4px;
}
QScrollBar::handle:horizontal{
    background:""" + T['border_hi'] + """;border-radius:4px;
}
QTabWidget::pane{
    border:1px solid """ + T['border'] + """;
    background:""" + T['bg_panel'] + """;
    border-radius:0 8px 8px 8px;
}
QTabBar::tab{
    background:""" + T['bg_header'] + """;
    color:""" + T['text_secondary'] + """;
    padding:7px 18px;
    border:1px solid """ + T['border'] + """;
    border-bottom:none;
    border-radius:8px 8px 0 0;
    margin-right:2px;
    font-size:12px;
}
QTabBar::tab:selected{
    background:""" + T['bg_panel'] + """;
    color:""" + T['accent'] + """;
    border-top:2px solid """ + T['accent'] + """;
    font-weight:bold;font-size:13px;
}
QStatusBar{
    background:""" + T['bg_header'] + """;
    color:""" + T['text_secondary'] + """;
    border-top:1px solid """ + T['border'] + """;
    font-size:12px;padding:3px 6px;
}
QToolBar{
    background:""" + T['bg_panel'] + """;
    border-bottom:1px solid """ + T['border'] + """;
    spacing:4px;padding:5px 10px;
}
QToolBar::separator{background:""" + T['border'] + """;width:1px;margin:4px 8px;}
QToolBar QToolButton{
    background-color:""" + T['accent'] + """;
    color:white;
    border:1px solid """ + T['accent_dark'] + """;
    border-radius:10px;
    padding:6px 14px;
    font-weight:700;
    font-size:13px;
    min-height:30px;
}
QToolBar QToolButton:hover{
    background-color:""" + T['accent_dark'] + """;
    border:1px solid """ + T['accent_dark'] + """;
}
QToolBar QToolButton:pressed{
    background-color:""" + T['accent_dark'] + """;
    padding-top:7px;   /* subtle "press" feel */
    padding-bottom:5px;
}
"""

# ─────────────────────────────────────────────────────────────────────────────
#  CUSTOM COMBOBOX  — draws its own crisp chevron arrow
# ─────────────────────────────────────────────────────────────────────────────

class ArrowComboBox(QComboBox):
    """QComboBox that paints a clean chevron arrow in the drop-down button."""

    def showPopup(self):
        super().showPopup()

        # On Windows, the popup can have native shadow/frames that QSS can't remove.
        # Force a square, no-shadow popup for a clean look.
        popup = self.view().window()
        if popup is None:
            return
        try:
            popup.setWindowFlag(Qt.FramelessWindowHint, True)
            popup.setWindowFlag(Qt.NoDropShadowWindowHint, True)
        except AttributeError:
            # Some Qt builds may not support NoDropShadowWindowHint; ignore gracefully.
            popup.setWindowFlag(Qt.FramelessWindowHint, True)

        popup.setAttribute(Qt.WA_TranslucentBackground, False)
        popup.setStyleSheet(
            f"background:{T['bg_panel']};"
            f"border:1px solid {T['border_hi']};"
            f"border-radius:0px;"
        )
        popup.show()

    def paintEvent(self, event):
        super().paintEvent(event)
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        # drop-down button area (right 32 px)
        btn_x = self.width() - 32
        btn_w = 32
        cx = btn_x + btn_w / 2.0
        cy = self.height() / 2.0

        # chevron  ▾  — three points
        aw = 8.0   # half-width of chevron
        ah = 5.0   # height of chevron
        pts = QPolygonF([
            QPointF(cx - aw, cy - ah / 2),
            QPointF(cx,      cy + ah / 2),
            QPointF(cx + aw, cy - ah / 2),
        ])
        pen = QPen(QColor(T['accent']), 2.0, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        p.setPen(pen)
        p.setBrush(Qt.NoBrush)
        p.drawPolyline(pts)
        p.end()

# ─────────────────────────────────────────────────────────────────────────────
#  APP ICON
# ─────────────────────────────────────────────────────────────────────────────

def make_app_icon() -> QIcon:
    # Prefer a local icon asset if present (portable across machines).
    # Drop a `wafer_icon` file next to this script (e.g. wafer_icon.ico/png/icns).
    base_dir = os.path.dirname(__file__)
    preferred_exts = ('.ico', '.png', '.icns', '.jpg', '.jpeg')
    try:
        for name in os.listdir(base_dir):
            low = name.lower()
            if low.startswith('wafer_icon') and low.endswith(preferred_exts):
                pth = os.path.join(base_dir, name)
                if os.path.isfile(pth):
                    return QIcon(pth)
    except OSError:
        pass

    icon = QIcon()
    for sz in (16, 24, 32, 48, 64, 128):
        pix = QPixmap(sz, sz)
        pix.fill(Qt.transparent)
        p = QPainter(pix)
        p.setRenderHint(QPainter.Antialiasing)
        cx = cy = sz / 2.0
        r  = sz / 2.0 - 1.0
        disc = QRadialGradient(cx - r*0.2, cy - r*0.25, r*1.35)
        disc.setColorAt(0, QColor("#2d5490"))
        disc.setColorAt(0.65, QColor("#1a3460"))
        disc.setColorAt(1.0, QColor("#0c1e3a"))
        p.setBrush(QBrush(disc))
        p.setPen(QPen(QColor("#5588cc"), max(1.0, sz/18.0)))
        p.drawEllipse(QPointF(cx, cy), r, r)
        nw = r*0.36; nh = max(2.0, sz*0.05)
        p.setCompositionMode(QPainter.CompositionMode_Clear)
        p.setPen(Qt.NoPen); p.setBrush(Qt.transparent)
        p.drawRect(QRectF(cx-nw/2, cy+r-nh, nw, nh+1))
        p.setCompositionMode(QPainter.CompositionMode_SourceOver)
        p.setPen(QPen(QColor("#5588cc"), max(1.0, sz/20.0)))
        p.drawLine(QPointF(cx-nw/2, cy+r-nh), QPointF(cx+nw/2, cy+r-nh))
        colors = [QColor("#3dba6f"), QColor("#3dba6f"), QColor("#8a98b0"),
                  QColor("#3dba6f"), QColor("#3dba6f"), QColor("#d95555"),
                  QColor("#8a98b0"), QColor("#3dba6f"), QColor("#3dba6f")]
        ge = r*1.08; cs = ge/3.0; gap = max(0.6, cs*0.12)
        gx0 = cx - ge/2; gy0 = cy - ge/2
        p.setPen(QPen(QColor("#0c1e3a"), max(0.5, gap*0.4)))
        for row in range(3):
            for col in range(3):
                rx = gx0 + col*cs + gap; ry = gy0 + row*cs + gap
                rw = rh = cs - gap*2
                if rw > 0:
                    p.setBrush(QBrush(colors[row*3+col]))
                    p.drawRoundedRect(QRectF(rx, ry, rw, rh),
                                      max(0.5, rw*0.18), max(0.5, rw*0.18))
        p.end()
        icon.addPixmap(pix)
    return icon

# ─────────────────────────────────────────────────────────────────────────────
#  KDF PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_kdf(filepath: str):
    header: dict = {}
    sites:  list = []
    mkeys:  set  = set()
    tkeys:  set  = set()

    with open(filepath, 'r', errors='replace') as fh:
        lines = [l.rstrip('\r\n') for l in fh]

    i = 0
    while i < len(lines):
        s = lines[i].strip()
        if s == '<EOH>':
            i += 1
            break
        if ',' in s:
            k, _, v = s.partition(',')
            header[k.strip()] = v.strip()
        i += 1

    if i < len(lines) and lines[i].strip() and not lines[i].strip().startswith('Site_'):
        i += 1

    current = None
    while i < len(lines):
        s = lines[i].strip(); i += 1
        if not s:
            continue
        if s == '<EOS>':
            if current is not None:
                sites.append(current)
            current = None
            continue
        if s.startswith('Site_'):
            if current is not None:
                sites.append(current)
            parts = s.split(',')
            name  = parts[0].strip()
            try:
                x = int(parts[1].strip())
                y = int(parts[2].strip())
            except (IndexError, ValueError):
                m = re.match(r'Site_([pn])(\d+)([pn])(\d+)', name)
                if m:
                    xs, xv, ys, yv = m.groups()
                    x = int(xv) * (1 if xs == 'p' else -1)
                    y = int(yv) * (1 if ys == 'p' else -1)
                else:
                    x = y = 0
            current = {'name': name, 'x': x, 'y': y, 'subsites': {}}
            continue
        if current is None:
            continue
        if '@' in s and ',' in s:
            kp, _, vs = s.partition(',')
            parts = kp.split('@')
            if len(parts) >= 3:
                param, test, sub_raw = parts[0].strip(), parts[1].strip(), parts[2].strip()
                try:
                    sub_num = int(sub_raw.split('#')[1])
                except (IndexError, ValueError):
                    sub_num = 1
                try:
                    value = float(vs.strip())
                except ValueError:
                    value = None
                mkey = f"{param}@{test}"
                mkeys.add(mkey); tkeys.add(test)
                if sub_num not in current['subsites']:
                    current['subsites'][sub_num] = {}
                current['subsites'][sub_num][mkey] = value

    if current is not None:
        sites.append(current)

    return header, sites, sorted(mkeys), sorted(tkeys)


def get_site_value(site: dict, mkey: str, subsite: int | None = None) -> float | None:
    if subsite is not None:
        v = site['subsites'].get(subsite, {}).get(mkey)
        return v if (v is not None and math.isfinite(v)) else None

    # Aggregate mode is only meaningful for single-design dies.
    values = {
        sn: sub_vals.get(mkey)
        for sn, sub_vals in site['subsites'].items()
    }
    finite = [v for v in values.values() if v is not None and math.isfinite(v)]
    if not finite:
        return None
    if len(finite) == 1:
        return finite[0]
    return None


def all_subsites(sites: list) -> list[int]:
    s = set()
    for site in sites:
        s.update(site['subsites'].keys())
    return sorted(s)


def si_fmt(v: float | None) -> str:
    if v is None or not math.isfinite(v):
        return 'N/A'
    if v == 0:
        return '0'

    prefixes = [
        (1e15,'P'), (1e12,'T'), (1e9,'G'),  (1e6,'M'),  (1e3,'k'),
        (1e0,''),   (1e-3,'m'),(1e-6,'µ'),  (1e-9,'n'), (1e-12,'p'),
        (1e-15,'f'),(1e-18,'a'),
    ]
    av = abs(v)
    best_factor, best_suffix = prefixes[-1]
    for factor, suffix in prefixes:
        if 0.1 <= av / factor < 1000.0:
            best_factor = factor; best_suffix = suffix; break

    scaled = v / best_factor
    if abs(scaled) >= 100:   text = f"{scaled:.1f}"
    elif abs(scaled) >= 10:  text = f"{scaled:.2f}"
    else:                    text = f"{scaled:.3f}"
    if '.' in text:
        text = text.rstrip('0').rstrip('.')
    return f"{text}{best_suffix}"

# ─────────────────────────────────────────────────────────────────────────────
#  WAFER CANVAS
# ─────────────────────────────────────────────────────────────────────────────

class WaferCanvas(QWidget):
    siteClicked = Signal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.sites         = []
        self.values        = {}
        self.low_limit     = None
        self.high_limit    = None
        self.prod_low_limit = None
        self.prod_high_limit = None
        self.show_prod_limits = False
        self.mkey          = ''
        self.selected_site = None
        self._hover        = None
        self._rects        = {}
        self._zoom         = 1.0
        self._continuous_mode = False

        self.setMinimumSize(400, 400)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setMouseTracking(True)

    def load(self, sites, values, lo, hi, mkey='', prod_lo=None, prod_hi=None, show_prod=False):
        self.sites = sites; self.values = values
        self.low_limit = lo; self.high_limit = hi; self.mkey = mkey
        self.prod_low_limit = prod_lo
        self.prod_high_limit = prod_hi
        self.show_prod_limits = bool(show_prod)
        self.selected_site = None; self._hover = None
        self.update()

    def set_continuous_mode(self, enabled: bool):
        self._continuous_mode = bool(enabled)
        self.update()

    def _value_range(self):
        vals = [v for v in self.values.values() if v is not None and math.isfinite(v)]
        if not vals:
            return None, None
        return min(vals), max(vals)

    @property
    def _limits_active(self):
        return self.low_limit is not None or self.high_limit is not None

    def _bounds(self):
        if not self.sites:
            return -1, 1, -1, 1
        xs = [s['x'] for s in self.sites]
        ys = [s['y'] for s in self.sites]
        return min(xs), max(xs), min(ys), max(ys)

    def _layout(self, x0, x1, y0, y1, w, h):
        n_cols = x1 - x0 + 1
        n_rows = y1 - y0 + 1

        cx = w / 2.0
        cy = h / 2.0
        radius_max = min(cx, w - cx, cy, h - cy) - 0.5
        disc_margin = 2.0
        radius = max(1.0, radius_max - disc_margin)

        cell_hi = min(w / max(n_cols, 1), h / max(n_rows, 1))
        cell_lo = 1.0

        def fits(cell: float) -> bool:
            mg = max(1.0, cell * 0.03)
            half_grid_w = cell * n_cols / 2.0 - mg
            half_grid_h = cell * n_rows / 2.0 - mg
            half_grid_w = max(0.0, half_grid_w)
            half_grid_h = max(0.0, half_grid_h)
            dist = math.hypot(half_grid_w, half_grid_h)
            return dist <= (radius - 0.3)

        for _ in range(24):
            cell_mid = (cell_lo + cell_hi) / 2.0
            if fits(cell_mid):
                cell_lo = cell_mid
            else:
                cell_hi = cell_mid

        cell = cell_lo
        ox = (w - cell * n_cols) / 2.0
        oy = (h - cell * n_rows) / 2.0
        return ox, oy, cell

    # ── zoom controls ────────────────────────────────────────────────────────
    def zoom_in(self):
        self.set_zoom(self._zoom * 1.15)

    def zoom_out(self):
        self.set_zoom(self._zoom / 1.15)

    def reset_zoom(self):
        self.set_zoom(1.0)

    def set_zoom(self, z: float):
        z = max(0.6, min(3.5, float(z)))
        if abs(z - self._zoom) < 1e-6:
            return
        self._zoom = z
        self.update()

    def _to_logical(self, pos: QPointF) -> QPointF:
        # Map widget coords -> logical coords used by _rects when zoomed.
        cx = self.width() / 2.0
        cy = self.height() / 2.0
        return QPointF((pos.x() - cx) / self._zoom + cx, (pos.y() - cy) / self._zoom + cy)

    def _die_color(self, name):
        v = self.values.get(name)
        if v is None:
            return QColor(T['nodata_bg']), QColor(T['nodata_fg']), QColor(T['nodata_border'])
        if self._continuous_mode:
            vmin, vmax = self._value_range()
            if vmin is None or vmax is None:
                return QColor(T['neutral_bg']), QColor(T['neutral_fg']), QColor(T['neutral_border'])
            if abs(vmax - vmin) < 1e-18:
                n = 0.5
            else:
                n = (v - vmin) / (vmax - vmin)
            # Saturated green -> yellow -> red ramp using the requested palette.
            green = QColor(PASS_COLOR)
            yellow = QColor(WARN_COLOR)
            red = QColor(FAIL_COLOR)
            if n <= 0.5:
                bg = _lerp_color(green, yellow, n * 2.0)
            else:
                bg = _lerp_color(yellow, red, (n - 0.5) * 2.0)
            fg = QColor(T['text_primary'])
            bc = bg.darker(120)
            return bg, fg, bc
        if not self._limits_active:
            return QColor(T['neutral_bg']), QColor(T['neutral_fg']), QColor(T['neutral_border'])
        lo, hi  = self.low_limit, self.high_limit
        spec_passed  = (lo is None or v >= lo) and (hi is None or v <= hi)
        if not spec_passed:
            return QColor(T['fail_bg']), QColor(T['fail_fg']), QColor(T['fail_border'])

        prod_active = self.show_prod_limits and (
            self.prod_low_limit is not None or self.prod_high_limit is not None
        )
        if prod_active:
            prod_passed = (
                (self.prod_low_limit is None or v >= self.prod_low_limit) and
                (self.prod_high_limit is None or v <= self.prod_high_limit)
            )
            if not prod_passed:
                return QColor(T['warn_bg']), QColor(T['warn_fg']), QColor(T['warn_border'])

        if spec_passed:
            return QColor(T['pass_bg']), QColor(T['pass_fg']), QColor(T['pass_border'])
        return QColor(T['neutral_bg']), QColor(T['neutral_fg']), QColor(T['neutral_border'])

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.setRenderHint(QPainter.TextAntialiasing)
        w, h = self.width(), self.height()

        p.fillRect(0, 0, w, h, QColor(T['bg_app']))
        p.setPen(QPen(QColor(T['border']), 1.0))
        for gx in range(0, w + 28, 28):
            for gy in range(0, h + 28, 28):
                p.drawPoint(gx, gy)

        if not self.sites:
            p.setPen(QColor(T['text_dim']))
            p.setFont(QFont('Segoe UI', 14))
            p.drawText(self.rect(), Qt.AlignCenter, 'Open a KDF file to begin')
            return

        # Zoom wafer content about the view center; draw legend unzoomed.
        cx_dev = w / 2.0
        cy_dev = h / 2.0
        p.save()
        p.translate(cx_dev, cy_dev)
        p.scale(self._zoom, self._zoom)
        p.translate(-cx_dev, -cy_dev)

        lw = w / self._zoom
        lh = h / self._zoom
        lx0 = cx_dev - lw / 2.0
        ly0 = cy_dev - lh / 2.0

        x0, x1, y0, y1 = self._bounds()
        ox, oy, cell    = self._layout(x0, x1, y0, y1, lw, lh)
        ox += lx0
        oy += ly0

        cx = cx_dev
        cy = cy_dev
        radius_max = min(lw / 2.0, lh / 2.0) - 0.5
        disc_margin = 2.0
        radius = max(1.0, radius_max - disc_margin)

        shad = QRadialGradient(cx+4, cy+4, radius+4)
        shad.setColorAt(0,   QColor(0, 0, 0, 24))
        shad.setColorAt(0.8, QColor(0, 0, 0,  8))
        shad.setColorAt(1.0, QColor(0, 0, 0,  0))
        p.setBrush(QBrush(shad)); p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(cx+4, cy+4), radius+4, radius+4)

        wg = QRadialGradient(cx - radius*0.15, cy - radius*0.2, radius*1.3)
        wg.setColorAt(0,    QColor('#ffffff'))
        wg.setColorAt(0.5,  QColor(T['wafer_bg']))
        wg.setColorAt(1.0,  QColor('#cfd9e8'))
        p.setBrush(QBrush(wg))
        p.setPen(QPen(QColor(T['wafer_edge']), 1.5))
        p.drawEllipse(QPointF(cx, cy), radius, radius)

        nw = radius * 0.22
        p.setPen(Qt.NoPen); p.setBrush(QColor(T['bg_app']))
        p.drawRect(QRectF(cx - nw/2, cy + radius - 6, nw, 6))
        p.setPen(QPen(QColor(T['wafer_edge']), 2))
        p.drawLine(QPointF(cx - nw/2, cy + radius - 4),
                   QPointF(cx + nw/2, cy + radius - 4))

        # die corner radius — scales with cell size, minimum 3 px
        die_radius = max(3.0, cell * 0.10)

        fs   = max(7, int(cell * 0.15))
        vfnt = QFont('Consolas', fs, QFont.Bold)
        cfnt = QFont('Consolas', max(6, fs - 2))
        self._rects = {}

        for site in self.sites:
            mg   = max(1.0, cell * 0.03)
            px_  = ox + (site['x'] - x0) * cell
            py_  = oy + (y1 - site['y']) * cell
            rect = QRectF(px_ + mg, py_ + mg, cell - 2*mg, cell - 2*mg)
            self._rects[site['name']] = rect

            bg, fg, bc = self._die_color(site['name'])
            is_sel = self.selected_site and site['name'] == self.selected_site['name']
            is_hov = self._hover         and site['name'] == self._hover['name']

            if cell > 30:
                p.setBrush(QColor(0,0,0,12)); p.setPen(Qt.NoPen)
                p.drawRoundedRect(
                    QRectF(rect.x()+2, rect.y()+2, rect.width(), rect.height()),
                    die_radius, die_radius)

            cg = QLinearGradient(rect.topLeft(), rect.bottomRight())
            # Keep the highlight subtle so the ramp stays saturated.
            cg.setColorAt(0, bg.lighter(104)); cg.setColorAt(1, bg)
            p.setBrush(QBrush(cg))
            p.setPen(QPen(QColor(T['selected']), 2.5) if is_sel
                     else QPen(QColor(T['hover_border']), 2) if is_hov
                     else QPen(bc, 1))
            p.drawRoundedRect(rect, die_radius, die_radius)

            v = self.values.get(site['name'])
            if v is not None and cell > 28:
                p.setFont(vfnt); p.setPen(fg)
                p.drawText(rect, Qt.AlignCenter, si_fmt(v))

            if cell > 58:
                p.setFont(cfnt); p.setPen(QColor(T['text_dim']))
                cr = QRectF(px_+mg+2, py_+mg+1, cell-2*mg-2, cell*0.28)
                p.drawText(cr, Qt.AlignLeft | Qt.AlignTop,
                           f"{site['x']},{site['y']}")

        p.restore()
        self._draw_legend(p, w, h)

    def _draw_legend(self, p, w, h):
        if self._continuous_mode:
            vals = [v for v in self.values.values() if v is not None and math.isfinite(v)]
            if not vals:
                return
            vmin = min(vals)
            vmax = max(vals)
            lx = 14
            ly = h - 76
            panel_w = 238
            panel_h = 54
            p.setBrush(QColor(255, 250, 242, 215))
            p.setPen(QPen(QColor(T['border']), 1))
            p.drawRoundedRect(QRectF(lx - 6, ly - 8, panel_w, panel_h), 8, 8)

            grad_rect = QRectF(lx, ly + 6, 180, 14)
            lg = QLinearGradient(grad_rect.left(), grad_rect.top(), grad_rect.right(), grad_rect.top())
            lg.setColorAt(0.0, QColor(PASS_COLOR))
            lg.setColorAt(0.5, QColor(WARN_COLOR))
            lg.setColorAt(1.0, QColor(FAIL_COLOR))
            p.setPen(QPen(QColor(T['border']), 1))
            p.setBrush(QBrush(lg))
            p.drawRoundedRect(grad_rect, 3, 3)
            p.setPen(QColor(T['text_secondary']))
            p.setFont(QFont('Consolas', 9))
            p.drawText(QRectF(lx, ly + 24, 80, 16), Qt.AlignLeft | Qt.AlignVCenter, si_fmt(vmin))
            p.drawText(QRectF(lx + 78, ly + 24, 24, 16), Qt.AlignCenter, 'mid')
            p.drawText(QRectF(lx + 100, ly + 24, 80, 16), Qt.AlignRight | Qt.AlignVCenter, si_fmt(vmax))
            p.setPen(QColor(T['text_primary']))
            p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(lx + 188, ly + 6, 42, 30), Qt.AlignCenter, 'Value\nramp')
            return

        prod_active = self.show_prod_limits and (
            self.prod_low_limit is not None or self.prod_high_limit is not None
        )
        if self._limits_active and prod_active:
            items = [
                (T['pass_bg'], T['pass_fg'], 'Pass (production)'),
                (T['warn_bg'], T['warn_fg'], 'Pass spec / Fail prod'),
                (T['fail_bg'], T['fail_fg'], 'Fail spec'),
                (T['nodata_bg'], T['nodata_fg'], 'No data'),
            ]
        elif self._limits_active:
            items = [
                (T['pass_bg'], T['pass_fg'], 'Pass'),
                (T['fail_bg'], T['fail_fg'], 'Fail'),
                (T['nodata_bg'], T['nodata_fg'], 'No data'),
            ]
        else:
            items = [
                (T['neutral_bg'], T['neutral_fg'], 'No limits set'),
                (T['nodata_bg'], T['nodata_fg'], 'No data'),
            ]

        bh = len(items) * 22 + 16
        lx = 14; ly = h - bh - 10
        # Warm translucent legend panel
        p.setBrush(QColor(255, 250, 242, 215))
        p.setPen(QPen(QColor(T['border']), 1))
        p.drawRoundedRect(QRectF(lx-6, ly-8, 178, bh), 8, 8)
        p.setFont(QFont('Segoe UI', 11))
        for bg_hex, fg_hex, label in items:
            p.setBrush(QBrush(QColor(bg_hex)))
            p.setPen(QPen(QColor(fg_hex), 1))
            p.drawRoundedRect(QRectF(lx, ly, 14, 14), 4, 4)
            p.setPen(QColor(T['text_primary']))
            p.drawText(int(lx+20), int(ly+11), label)
            ly += 22

    def mouseMoveEvent(self, e):
        pos = self._to_logical(QPointF(e.position())); self._hover = None
        for site in self.sites:
            r = self._rects.get(site['name'])
            if r and r.contains(pos):
                self._hover = site
                self.setCursor(Qt.PointingHandCursor)
                break
        else:
            self.setCursor(Qt.ArrowCursor)
        self.update()

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            pos = self._to_logical(QPointF(e.position()))
            for site in self.sites:
                r = self._rects.get(site['name'])
                if r and r.contains(pos):
                    self.selected_site = site
                    self.siteClicked.emit(site)
                    self.update()
                    return

    def leaveEvent(self, _e):
        self._hover = None; self.update()

# ─────────────────────────────────────────────────────────────────────────────
#  DIE DETAIL PANEL
# ─────────────────────────────────────────────────────────────────────────────

class _DesignTintDelegate(QStyledItemDelegate):
    def __init__(self, colors: list[QColor], role_design: int, parent=None):
        super().__init__(parent)
        self._colors = colors
        self._role_design = role_design

    def _color_for_design(self, design_num: int) -> QColor:
        return self._colors[(design_num - 1) % len(self._colors)]

    def paint(self, painter: QPainter, option, index):
        design_num = index.data(self._role_design)
        if isinstance(design_num, int):
            base = QColor(self._color_for_design(design_num))

            # Professional look: very subtle row wash + a clean left accent stripe.
            c = QColor(base)
            c.setAlpha(18)
            painter.save()
            painter.fillRect(option.rect, c)

            # Accent stripe only on first column to avoid looking "busy".
            if index.column() == 0:
                stripe = QColor(base)
                stripe.setAlpha(170)
                r = option.rect
                painter.fillRect(QRect(r.x(), r.y(), 4, r.height()), stripe)
            painter.restore()

        super().paint(painter, option, index)


class SiteDetailPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        lo = QVBoxLayout(self); lo.setContentsMargins(8, 8, 8, 8); lo.setSpacing(8)

        self.title = QLabel('Click a die to inspect')
        self.title.setWordWrap(True)
        self.title.setStyleSheet(
            f'font-weight:bold;font-size:13px;color:{T["accent_dark"]};'
            f'padding:5px 8px;background:{T["accent_dim"]};'
            f'border-radius:8px;border-left:3px solid {T["accent"]};')
        lo.addWidget(self.title)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(['Measurement', 'Design #', 'Value'])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        # Alternating row colors fight with design tinting; keep the look clean.
        self.table.setAlternatingRowColors(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        # Prevent global QSS hover/selection from wiping our design tint.
        self.table.setStyleSheet(
            f"QTableWidget::item:hover{{background-color: rgba(0,0,0,0); border:1px solid {T['border']};}}"
            f"QTableWidget::item:selected{{background-color: rgba(0,0,0,0); border:1px solid {T['accent']}; color:{T['accent_dark']};}}"
        )
        lo.addWidget(self.table)

        self._design_colors = [
            QColor("#1565c0"),  # blue
            QColor("#2e7d32"),  # green
            QColor("#6a1b9a"),  # purple
            QColor("#c62828"),  # red
            QColor("#ef6c00"),  # orange
            QColor("#00838f"),  # teal
            QColor("#4e342e"),  # brown
            QColor("#455a64"),  # blue-grey
        ]
        self._role_design = int(Qt.UserRole) + 1
        self.table.setItemDelegate(
            _DesignTintDelegate(self._design_colors, self._role_design, self.table)
        )

    def _design_color(self, design_num: int) -> QColor:
        # stable color assignment per design number
        return self._design_colors[(design_num - 1) % len(self._design_colors)]

    def show_site(self, site: dict):
        self.title.setText(
            f'  {site["name"]}   ·   X = {site["x"]},  Y = {site["y"]}')

        rows = []
        for sub_num in sorted(site['subsites']):
            for mkey, val in sorted(site['subsites'][sub_num].items()):
                rows.append((mkey, sub_num, val))

        self.table.setRowCount(len(rows))
        for i, (mkey, sub, val) in enumerate(rows):
            multi = len(site.get('subsites', {})) > 1
            design_num = int(sub) if multi else None

            mi = QTableWidgetItem(mkey)
            mi.setFont(QFont('Segoe UI', 12))
            if design_num is not None:
                mi.setData(self._role_design, design_num)
            self.table.setItem(i, 0, mi)

            si = QTableWidgetItem(str(sub))
            si.setTextAlignment(Qt.AlignCenter)
            si.setForeground(QColor(T['text_secondary']))
            si.setFont(QFont('Segoe UI', 12))
            if design_num is not None:
                si.setData(self._role_design, design_num)
                si.setForeground(QColor(T['text_primary']))

            self.table.setItem(i, 1, si)

            display = si_fmt(val) if (val is not None and math.isfinite(val)) else 'N/A'
            vi = QTableWidgetItem(display)
            vi.setForeground(QColor(T['accent_dark']))
            vi.setFont(QFont('Consolas', 12, QFont.Bold))

            if design_num is not None:
                vi.setData(self._role_design, design_num)

            self.table.setItem(i, 2, vi)

# ─────────────────────────────────────────────────────────────────────────────
#  STATISTICS PANEL
# ─────────────────────────────────────────────────────────────────────────────

class StatsPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        lo = QVBoxLayout(self); lo.setContentsMargins(8, 8, 8, 8); lo.setSpacing(6)
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(['Statistic', 'Value'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        lo.addWidget(self.table)

    def update_stats(self, values_dict: dict, lo, hi):
        vals = [v for v in values_dict.values()
                if v is not None and math.isfinite(v)]
        if not vals:
            self.table.setRowCount(0)
            return

        n      = len(vals)
        mean_v = statistics.mean(vals)
        std_v  = statistics.pstdev(vals)
        med_v  = statistics.median(vals)
        min_v  = min(vals)
        max_v  = max(vals)
        passed = sum(1 for v in vals
                     if (lo is None or v >= lo) and (hi is None or v <= hi))
        yld    = passed / n * 100

        rows = [
            ('Count',   str(n)),
            ('Mean',    si_fmt(mean_v)),
            ('Std Dev', si_fmt(std_v)),
            ('Min',     si_fmt(min_v)),
            ('Max',     si_fmt(max_v)),
            ('Median',  si_fmt(med_v)),
            ('3σ',      f'{si_fmt(mean_v - 3*std_v)}  →  {si_fmt(mean_v + 3*std_v)}'),
            ('Pass',    str(passed)),
            ('Fail',    str(n - passed)),
            ('Yield',   f'{yld:.1f}%'),
        ]
        self.table.setRowCount(len(rows))
        for i, (k, v) in enumerate(rows):
            ki = QTableWidgetItem(k)
            ki.setForeground(QColor(T['text_secondary']))
            ki.setFont(QFont('Segoe UI', 12))
            self.table.setItem(i, 0, ki)

            vi = QTableWidgetItem(v)
            vi.setFont(QFont('Consolas', 12))
            if k == 'Yield':
                col = (T['pass_fg'] if yld >= 90
                       else T['fail_fg'] if yld < 70
                       else T['warn'])
                vi.setForeground(QColor(col))
                vi.setFont(QFont('Segoe UI', 13, QFont.Bold))
            elif k == 'Pass': vi.setForeground(QColor(T['pass_fg']))
            elif k == 'Fail': vi.setForeground(QColor(T['fail_fg']))
            else:             vi.setForeground(QColor(T['text_primary']))
            self.table.setItem(i, 1, vi)


class HistogramPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._values = []
        self._lo = None
        self._hi = None
        self._bins = 20
        self.setMinimumHeight(180)

    def set_data(self, values: list[float], lo, hi):
        self._values = [v for v in values if v is not None and math.isfinite(v)]
        self._lo = lo
        self._hi = hi
        self.update()

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.fillRect(self.rect(), QColor(T['bg_panel']))
        if not self._values:
            p.setPen(QColor(T['text_dim']))
            p.drawText(self.rect(), Qt.AlignCenter, 'No data for histogram')
            return
        vals = self._values
        vmin, vmax = min(vals), max(vals)
        if abs(vmax - vmin) < 1e-18:
            vmin -= 0.5
            vmax += 0.5
        bins = [0] * self._bins
        span = vmax - vmin
        for v in vals:
            idx = int((v - vmin) / span * self._bins)
            idx = max(0, min(self._bins - 1, idx))
            bins[idx] += 1
        max_bin = max(bins) if bins else 1

        chart = QRectF(48, 18, max(80, self.width() - 64), max(120, self.height() - 56))
        p.setPen(QPen(QColor(T['border']), 1))
        p.drawRect(chart)
        bw = chart.width() / self._bins
        for i, cnt in enumerate(bins):
            h = (cnt / max_bin) * (chart.height() - 4) if max_bin else 0
            r = QRectF(chart.left() + i * bw + 1, chart.bottom() - h - 1, max(1.0, bw - 2), h)
            p.fillRect(r, QColor(T['accent_dim']))
            p.setPen(QColor(T['accent']))
            p.drawRect(r)

        mean_v = statistics.mean(vals)
        std_v = statistics.pstdev(vals)

        def x_for(v):
            return chart.left() + ((v - vmin) / (vmax - vmin)) * chart.width()

        markers = [
            ('Mean', mean_v, T['accent_dark'], Qt.SolidLine),
        ]
        for _label, mark, col, style in markers:
            if mark is None:
                continue
            x = x_for(mark)
            p.setPen(QPen(QColor(col), 1.5, style))
            p.drawLine(QPointF(x, chart.top()), QPointF(x, chart.bottom()))

        # Bell curve overlay (normal approximation) to complement Cp/Cpk context.
        if std_v > 0:
            span = vmax - vmin
            if span > 0 and max_bin > 0:
                bin_width = span / self._bins
                denom = std_v * math.sqrt(2.0 * math.pi)
                peak_pdf = 1.0 / denom
                expected_at_mean = peak_pdf * bin_width * len(vals)
                scale = (max_bin / expected_at_mean) if expected_at_mean > 0 else 1.0

                steps = 140
                curve = QPolygonF()
                for j in range(steps + 1):
                    x_val = vmin + (j / steps) * span
                    pdf = math.exp(-((x_val - mean_v) ** 2) / (2.0 * std_v * std_v)) / denom
                    expected = pdf * bin_width * len(vals) * scale
                    frac_h = max(0.0, min(1.0, expected / max_bin))
                    y = (chart.bottom() - 1) - frac_h * (chart.height() - 4)
                    curve.append(QPointF(x_for(x_val), y))

                p.setPen(QPen(QColor(T['accent_dark']), 2))
                p.drawPolyline(curve)

        # Axis labels and tick annotations.
        p.setPen(QColor(T['text_secondary']))
        font_ticks = QFont('Consolas', 9)
        p.setFont(font_ticks)
        p.drawText(QRectF(chart.left(), chart.bottom() + 2, 80, 14), si_fmt(vmin))
        p.drawText(QRectF(chart.right() - 80, chart.bottom() + 2, 80, 14), Qt.AlignRight, si_fmt(vmax))

        # X axis name: measurement value.
        p.setFont(QFont('Segoe UI', 9))
        p.drawText(QRectF(chart.left(), chart.bottom() + 18, chart.width(), 16),
                   Qt.AlignCenter, 'Measurement value')

        # Y axis name: count, rotated on the left.
        p.save()
        p.translate(chart.left() - 26, chart.top() + chart.height() / 2)
        p.rotate(-90)
        p.drawText(QRectF(-chart.height() / 2, -8, chart.height(), 16),
                   Qt.AlignCenter, 'Count')
        p.restore()

class YieldDonutPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        lo = QVBoxLayout(self); lo.setContentsMargins(0, 0, 0, 0); lo.setSpacing(6)
        self.summary = QLabel('Quick visual summary of pass and fail counts.')
        self.summary.setWordWrap(True)
        self.summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        self.summary.setVisible(False)
        lo.addWidget(self.summary)
        self._pass = 0
        self._fail = 0
        self._warn = 0
        self.setMinimumHeight(210)

    def set_data(self, passed: int, failed: int, warned: int = 0):
        self._pass = max(0, int(passed))
        self._fail = max(0, int(failed))
        self._warn = max(0, int(warned))
        self.update()

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.fillRect(self.rect(), QColor(T['bg_panel']))

        total = self._pass + self._fail
        if total <= 0:
            p.setPen(QColor(T['text_dim']))
            p.drawText(self.rect(), Qt.AlignCenter, 'No pass/fail data')
            return

        body = self.rect().adjusted(8, 32, -8, -8)
        donut_size = min(body.width() * 0.62, body.height() - 8)
        donut_size = max(96.0, donut_size)
        cx = body.left() + donut_size / 2.0 + 6
        cy = body.center().y()
        outer = QRectF(cx - donut_size / 2.0, cy - donut_size / 2.0, donut_size, donut_size)
        inner_margin = donut_size * 0.24
        start = 90 * 16
        pass_span = int(-360 * 16 * (self._pass / total))
        fail_span = -360 * 16 - pass_span

        p.setPen(Qt.NoPen)
        p.setBrush(QColor(PASS_COLOR))
        p.drawPie(outer, start, pass_span)
        p.setBrush(QColor(FAIL_COLOR))
        p.drawPie(outer, start + pass_span, fail_span)
        p.setBrush(QColor(T['bg_panel']))
        p.drawEllipse(outer.adjusted(inner_margin, inner_margin, -inner_margin, -inner_margin))

        p.setPen(QColor(T['text_primary']))
        p.setFont(QFont('Segoe UI', 16, QFont.Bold))
        p.drawText(outer, Qt.AlignCenter, f'{(self._pass / total) * 100.0:.0f}%')

        legend_x = outer.right() + 14
        legend_y = outer.top() + 18
        p.setFont(QFont('Segoe UI', 10))
        for idx, (label, value, color) in enumerate((
            ('Pass', self._pass, PASS_COLOR),
            ('Fail', self._fail, FAIL_COLOR),
            ('Near limit', self._warn, WARN_COLOR),
        )):
            if idx == 2 and value <= 0:
                continue
            y = legend_y + idx * 28
            p.setBrush(QColor(color))
            p.setPen(Qt.NoPen)
            p.drawEllipse(QRectF(legend_x, y, 10, 10))
            p.setPen(QColor(T['text_primary']))
            p.drawText(QRectF(legend_x + 16, y - 5, max(80, body.right() - legend_x - 12), 20),
                       Qt.AlignVCenter, f'{label}: {value}')


class MiniHeatmapPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._points = []
        self._vmin = None
        self._vmax = None
        self.setMinimumHeight(150)

    def set_data(self, points: list[dict], vmin=None, vmax=None):
        self._points = points
        self._vmin = vmin
        self._vmax = vmax
        self.update()

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.fillRect(self.rect(), QColor(T['bg_panel']))
        if not self._points:
            p.setPen(QColor(T['text_dim']))
            p.drawText(self.rect(), Qt.AlignCenter, 'No mini heatmap data')
            return

        xs = [d['x'] for d in self._points]
        ys = [d['y'] for d in self._points]
        x0, x1, y0, y1 = min(xs), max(xs), min(ys), max(ys)
        n_cols = x1 - x0 + 1
        n_rows = y1 - y0 + 1
        body = self.rect().adjusted(8, 8, -8, -8)
        cell = min(body.width() / max(1, n_cols), body.height() / max(1, n_rows))
        ox = body.left() + (body.width() - cell * n_cols) / 2.0
        oy = body.top() + (body.height() - cell * n_rows) / 2.0

        for d in self._points:
            value = d.get('value')
            if value is not None and self._vmin is not None and self._vmax is not None:
                if abs(self._vmax - self._vmin) < 1e-18:
                    n = 0.5
                else:
                    n = (value - self._vmin) / (self._vmax - self._vmin)
                n = max(0.0, min(1.0, n))
                if n <= 0.5:
                    bg = _lerp_color(QColor(PASS_COLOR), QColor(WARN_COLOR), n * 2.0)
                else:
                    bg = _lerp_color(QColor(WARN_COLOR), QColor(FAIL_COLOR), (n - 0.5) * 2.0)
            else:
                bg = QColor(T['nodata_bg'])
            x = ox + (d['x'] - x0) * cell
            y = oy + (y1 - d['y']) * cell
            rect = QRectF(x + 1, y + 1, max(2.0, cell - 2), max(2.0, cell - 2))
            p.setBrush(bg)
            p.setPen(QPen(QColor(T['border']), 0.8))
            p.drawRoundedRect(rect, 2.5, 2.5)


class ScatterPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._points = []
        self._xlabel = ''
        self._ylabel = ''
        self.setMinimumHeight(240)

    def set_data(self, points: list[tuple[float, float, bool]], xlabel: str, ylabel: str):
        self._points = points
        self._xlabel = xlabel
        self._ylabel = ylabel
        self.update()

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.fillRect(self.rect(), QColor(T['bg_panel']))
        chart = QRectF(44, 14, max(80, self.width() - 58), max(140, self.height() - 44))
        p.setPen(QPen(QColor(T['border']), 1))
        p.drawRect(chart)
        if not self._points:
            p.setPen(QColor(T['text_dim']))
            p.drawText(self.rect(), Qt.AlignCenter, 'No paired values for scatter plot')
            return
        xs = [pt[0] for pt in self._points]
        ys = [pt[1] for pt in self._points]
        xmin, xmax = min(xs), max(xs)
        ymin, ymax = min(ys), max(ys)
        if abs(xmax - xmin) < 1e-18:
            xmin -= 0.5; xmax += 0.5
        if abs(ymax - ymin) < 1e-18:
            ymin -= 0.5; ymax += 0.5

        def px(v):
            return chart.left() + (v - xmin) / (xmax - xmin) * chart.width()

        def py(v):
            return chart.bottom() - (v - ymin) / (ymax - ymin) * chart.height()

        for x, y, passed in self._points:
            col = QColor(T['pass_fg']) if passed else QColor(T['fail_fg'])
            p.setBrush(col)
            p.setPen(Qt.NoPen)
            p.drawEllipse(QPointF(px(x), py(y)), 3.2, 3.2)

        p.setPen(QColor(T['text_secondary']))
        p.setFont(QFont('Segoe UI', 9))
        p.drawText(QRectF(chart.left(), chart.bottom() + 3, chart.width(), 16), Qt.AlignCenter, self._xlabel)
        p.save()
        p.translate(12, chart.top() + chart.height() / 2)
        p.rotate(-90)
        p.drawText(QRectF(-chart.height() / 2, -12, chart.height(), 16), Qt.AlignCenter, self._ylabel)
        p.restore()


class YieldTrendPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._points = []
        self.setMinimumHeight(200)

    def set_data(self, points: list[tuple[str, float]]):
        self._points = points
        self.update()

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.fillRect(self.rect(), QColor(T['bg_panel']))
        # Leave room for axis title text at the bottom/left.
        chart = QRectF(36, 16, max(100, self.width() - 48), max(90, self.height() - 58))
        p.setPen(QPen(QColor(T['border']), 1))
        p.drawRect(chart)

        # Axis names for the lot trend view.
        p.setPen(QColor(T['text_secondary']))
        p.setFont(QFont('Segoe UI', 9))
        p.drawText(QRectF(chart.left(), chart.bottom() + 6, chart.width(), 16),
                   Qt.AlignCenter, 'Wafer sequence')
        p.save()
        p.translate(chart.left() - 26, chart.top() + chart.height() / 2)
        p.rotate(-90)
        p.drawText(QRectF(-chart.height() / 2, -8, chart.height(), 16),
                   Qt.AlignCenter, 'Yield (%)')
        p.restore()

        if len(self._points) < 2:
            p.setPen(QColor(T['text_dim']))
            p.drawText(chart, Qt.AlignCenter, 'Load batch and select measurement to view lot trend')
            return
        ys = [y for _, y in self._points]
        ymin, ymax = min(ys), max(ys)
        if abs(ymax - ymin) < 1e-9:
            ymin = max(0.0, ymin - 1.0)
            ymax = min(100.0, ymax + 1.0)
        pad = max(2.0, (ymax - ymin) * 0.08)
        ymin = max(0.0, ymin - pad)
        ymax = min(100.0, ymax + pad)

        poly = QPolygonF()
        n = len(self._points)
        for i, (_name, y) in enumerate(self._points):
            x = chart.left() + (i / (n - 1)) * chart.width()
            py = chart.bottom() - (y - ymin) / (ymax - ymin) * chart.height()
            poly.append(QPointF(x, py))
        p.setPen(QPen(QColor(T['accent']), 2))
        p.drawPolyline(poly)
        p.setBrush(QColor(T['accent_dark']))
        p.setPen(Qt.NoPen)
        for pt in poly:
            p.drawEllipse(pt, 2.8, 2.8)


class BatchFailHeatmapPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._points = []
        self.setMinimumHeight(320)

    def set_data(self, points: list[dict]):
        self._points = points
        self.update()

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.fillRect(self.rect(), QColor(T['bg_panel']))
        if not self._points:
            p.setPen(QColor(T['text_dim']))
            p.drawText(self.rect(), Qt.AlignCenter, 'No common fail-site data available')
            return
        xs = [d['x'] for d in self._points]
        ys = [d['y'] for d in self._points]
        x0, x1, y0, y1 = min(xs), max(xs), min(ys), max(ys)
        n_cols = x1 - x0 + 1
        n_rows = y1 - y0 + 1
        w = self.width() - 20
        h = self.height() - 20
        cell = min(w / max(1, n_cols), h / max(1, n_rows))
        ox = (self.width() - cell * n_cols) / 2.0
        oy = (self.height() - cell * n_rows) / 2.0
        for d in self._points:
            frac = d['fail_frac']
            # Saturated green -> yellow -> red ramp for fail frequency.
            green = QColor(PASS_COLOR)
            yellow = QColor(WARN_COLOR)
            red = QColor(FAIL_COLOR)
            if frac <= 0.0:
                bg = green
            elif frac >= 1.0:
                bg = red
            elif frac <= 0.5:
                bg = _lerp_color(green, yellow, frac * 2.0)
            else:
                bg = _lerp_color(yellow, red, (frac - 0.5) * 2.0)
            x = ox + (d['x'] - x0) * cell
            y = oy + (y1 - d['y']) * cell
            rect = QRectF(x + 1, y + 1, max(2.0, cell - 2), max(2.0, cell - 2))
            p.setBrush(bg)
            p.setPen(QPen(QColor(T['border']), 1))
            p.drawRoundedRect(rect, 3, 3)

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN WINDOW
# ─────────────────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Wafer Map Viewer')
        self.resize(1300, 840)
        self.setMinimumSize(1200, 760)
        self.setStyleSheet(SS)
        self.setWindowIcon(make_app_icon())

        self._header:       dict      = {}
        self._sites:        list      = []
        self._mkeys:        list[str] = []
        self._limits:       dict      = {}
        # When opening a wafer from "Batch Analysis" we must not wipe the batch-applied
        # limits; they're stored in `self._limits`.
        self._preserve_limits_on_next_load: bool = False
        self._current_mkey: str | None = None
        self._filepath:     str | None = None
        self._all_subsites: list[int]  = []
        self._current_sub:  int | None = None
        self._use_prod_limits: bool    = False
        self._batch_records: list[dict] = []
        self._batch_dir: str | None = None
        self._batch_rows: list[dict] = []
        self._raw_data_path: str | None = None

        self._build_ui()
        self._update_ui_state()
        # Keep startup behavior consistent across Windows versions.
        self.setWindowState(self.windowState() | Qt.WindowMaximized)

    def _build_ui(self):
        tb = QToolBar('Main', self)
        tb.setIconSize(QSize(18, 18)); tb.setMovable(False)
        self.addToolBar(tb)

        # Create early so toolbar actions can reference it.
        self.canvas = WaferCanvas()

        open_act = QAction('  Open KDF…', self)
        open_act.triggered.connect(self.open_file)
        tb.addAction(open_act)
        tb.addSeparator()

        reset_act = QAction('  Reset', self)
        reset_act.triggered.connect(self.reset_all)
        tb.addAction(reset_act)
        tb.addSeparator()

        exp_act = QAction('  Export Map…', self)
        exp_act.triggered.connect(self.export_map)
        tb.addAction(exp_act)
        tb.addSeparator()

        exp_xlsx_act = QAction('  Export Excel Map…', self)
        exp_xlsx_act.triggered.connect(self.export_map_excel)
        tb.addAction(exp_xlsx_act)
        tb.addSeparator()

        zoom_out = QAction('  −', self)
        zoom_out.setToolTip('Zoom out')
        zoom_out.triggered.connect(self.canvas.zoom_out)
        tb.addAction(zoom_out)

        zoom_in = QAction('  +', self)
        zoom_in.setToolTip('Zoom in')
        zoom_in.triggered.connect(self.canvas.zoom_in)
        tb.addAction(zoom_in)

        zoom_reset = QAction('  Reset Zoom', self)
        zoom_reset.setToolTip('Reset zoom')
        zoom_reset.triggered.connect(self.canvas.reset_zoom)
        tb.addAction(zoom_reset)
        tb.addSeparator()

        sp = QWidget(); sp.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        tb.addWidget(sp)
        self.lbl_file = QLabel('No file loaded  ')
        self.lbl_file.setStyleSheet(
            f'color:{T["accent_dark"]};'
            f'font-size:12px;'
            f'font-weight:700;'
            f'padding:4px 10px;'
            f'background:{T["accent_dim"]};'
            f'border:1px solid {T["border_hi"]};'
            f'border-radius:10px;'
            f'min-height:22px;'
        )
        tb.addWidget(self.lbl_file)

        cw = QWidget(); self.setCentralWidget(cw)
        root = QVBoxLayout(cw); root.setSpacing(8); root.setContentsMargins(10, 10, 10, 10)
        self.main_tabs = QTabWidget()
        root.addWidget(self.main_tabs)

        wafer_page = QWidget()
        mh = QHBoxLayout(wafer_page); mh.setSpacing(10); mh.setContentsMargins(0, 0, 0, 0)

        # ── left panel ────────────────────────────────────────────────────────
        left = QWidget(); left.setFixedWidth(320)
        lv = QVBoxLayout(left); lv.setSpacing(12); lv.setContentsMargins(0, 0, 0, 0)

        ib = QGroupBox('File Information')
        iform = QFormLayout(ib); iform.setSpacing(8)
        iform.setLabelAlignment(Qt.AlignRight)

        def mkval():
            l = QLabel('—')
            l.setStyleSheet(
                f'color:{T["text_primary"]};font-weight:700;font-size:14px;')
            return l

        self.lbl_lot   = mkval(); self.lbl_sys   = mkval()
        self.lbl_stt   = mkval(); self.lbl_cnt   = mkval()
        self.lbl_tests = mkval(); self.lbl_dsn   = mkval()

        for lbl_txt, wid in [
            ('Lot',     self.lbl_lot),
            ('System',  self.lbl_sys),
            ('Start',   self.lbl_stt),
            ('Sites',   self.lbl_cnt),
            ('Tests',   self.lbl_tests),
            ('Designs', self.lbl_dsn),
        ]:
            kl = QLabel(lbl_txt)
            kl.setStyleSheet(f'color:{T["text_secondary"]};font-size:13px;')
            iform.addRow(kl, wid)
        lv.addWidget(ib)

        self.left_batch_btn = QPushButton('Load Batch Folder…')
        self.left_batch_btn.setObjectName('primary')
        self.left_batch_btn.setMinimumHeight(34)
        self.left_batch_btn.clicked.connect(self.open_batch_folder)
        lv.addWidget(self.left_batch_btn)

        self.left_reset_btn = QPushButton('Reset All')
        self.left_reset_btn.clicked.connect(self.reset_all)
        lv.addWidget(self.left_reset_btn)

        # design selector — uses ArrowComboBox
        db = QGroupBox('Design')
        dv = QVBoxLayout(db); dv.setSpacing(8)
        desc = QLabel(
            'Each die may contain multiple designs.\n'
            'Select which design to display.')
        desc.setWordWrap(True)
        desc.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        dv.addWidget(desc)
        self.design_combo = ArrowComboBox()
        self.design_combo.setMinimumHeight(36)
        self.design_combo.currentIndexChanged.connect(self._on_design_changed)
        dv.addWidget(self.design_combo)
        lv.addWidget(db)

        # measurement selector — uses ArrowComboBox
        mb = QGroupBox('Measurement')
        mv = QVBoxLayout(mb); mv.setSpacing(8)
        self.mkey_combo = ArrowComboBox()
        self.mkey_combo.setMinimumHeight(36)
        self.mkey_combo.currentTextChanged.connect(self._on_mkey_changed)
        mv.addWidget(self.mkey_combo)
        lv.addWidget(mb)

        # pass/fail limits
        lb = QGroupBox('Pass / Fail Limits')
        lbv = QVBoxLayout(lb); lbv.setSpacing(8)
        for lbl_txt, attr in [('Low', 'low_edit'), ('High', 'high_edit')]:
            row = QHBoxLayout()
            rl = QLabel(lbl_txt)
            rl.setStyleSheet(
                f'color:{T["text_secondary"]};font-size:12px;min-width:36px;')
            row.addWidget(rl)
            en = QLineEdit(); en.setPlaceholderText('no limit')
            setattr(self, attr, en); row.addWidget(en); lbv.addLayout(row)

        self.prod_toggle = QCheckBox('Production limits')
        self.prod_toggle.setObjectName('prodToggle')
        self.prod_toggle.setStyleSheet(
            f"QCheckBox#prodToggle{{background:transparent;border:none;padding:2px 0;}}"
            f"QCheckBox#prodToggle::indicator{{"
            f"width:14px;height:14px;border:1px solid {T['border_hi']};"
            f"border-radius:3px;background:{T['bg_panel']};"
            f"}}"
            f"QCheckBox#prodToggle::indicator:checked{{"
            f"background:{T['accent']};border:1px solid {T['accent_dark']};"
            f"}}"
        )
        self.prod_toggle.toggled.connect(self._on_prod_toggle)
        lbv.addWidget(self.prod_toggle)

        self.prod_limits_wrap = QWidget()
        self.prod_limits_wrap.setStyleSheet("background:transparent;border:none;")
        plv = QVBoxLayout(self.prod_limits_wrap)
        plv.setContentsMargins(0, 0, 0, 0)
        plv.setSpacing(6)
        for lbl_txt, attr in [('Prod Low', 'prod_low_edit'), ('Prod High', 'prod_high_edit')]:
            row = QHBoxLayout()
            rl = QLabel(lbl_txt)
            rl.setStyleSheet(
                f'color:{T["text_secondary"]};font-size:12px;min-width:64px;')
            row.addWidget(rl)
            en = QLineEdit(); en.setPlaceholderText('no production limit')
            en.setStyleSheet(
                f"QLineEdit{{background:{T['bg_panel']};border:1px solid {T['border']};border-radius:8px;padding:5px 10px;}}"
                f"QLineEdit:focus{{border:1px solid {T['accent']};}}"
            )
            setattr(self, attr, en)
            row.addWidget(en)
            plv.addLayout(row)
        self.prod_limits_wrap.setVisible(False)
        lbv.addWidget(self.prod_limits_wrap)

        br = QHBoxLayout(); br.setSpacing(6)
        ab = QPushButton('Apply'); ab.setObjectName('primary')
        ab.clicked.connect(self._apply_limits)
        cb = QPushButton('Clear'); cb.clicked.connect(self._clear_limits)
        br.addWidget(ab); br.addWidget(cb); lbv.addLayout(br)
        lv.addWidget(lb)

        lv.addStretch()
        mh.addWidget(left)

        # ── centre canvas ─────────────────────────────────────────────────────
        canvas_wrap = QWidget()
        canvas_wrap.setStyleSheet(
            f'background:{T["bg_panel"]};border:1px solid {T["border"]};'
            f'border-radius:12px;')
        cl = QVBoxLayout(canvas_wrap); cl.setContentsMargins(4, 4, 4, 4)
        self.canvas.siteClicked.connect(self._on_die_clicked)
        cl.addWidget(self.canvas)
        mh.addWidget(canvas_wrap, stretch=3)

        # ── right panel ───────────────────────────────────────────────────────
        right = QWidget(); right.setFixedWidth(310)
        rv = QVBoxLayout(right); rv.setContentsMargins(0, 0, 0, 0)
        tabs = QTabWidget(); rv.addWidget(tabs)

        self.detail_panel = SiteDetailPanel()
        self.stats_panel = StatsPanel()
        self.analytics_panel = QWidget()
        av = QVBoxLayout(self.analytics_panel); av.setContentsMargins(8, 8, 8, 8); av.setSpacing(8)

        top_row = QHBoxLayout(); top_row.setSpacing(6)
        self.continuous_heatmap_toggle = QCheckBox('Continuous heatmap')
        self.continuous_heatmap_toggle.setToolTip('Toggle between discrete pass/fail coloring and continuous value heatmap.')
        self.continuous_heatmap_toggle.toggled.connect(self._on_continuous_heatmap_toggled)
        top_row.addWidget(self.continuous_heatmap_toggle)
        self.cpk_label = QLabel('Cp/Cpk: N/A')
        self.cpk_label.setWordWrap(True)
        self.cpk_label.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        top_row.addWidget(self.cpk_label, stretch=1)
        av.addLayout(top_row)

        dist_box = QGroupBox('Distribution')
        dv = QVBoxLayout(dist_box); dv.setContentsMargins(8, 8, 8, 8); dv.setSpacing(6)
        dist_summary = QLabel('Histogram with normal-fit overlay for the selected measurement.')
        dist_summary.setWordWrap(True)
        dist_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        dv.addWidget(dist_summary)
        self.hist_panel = HistogramPanel()
        self.hist_panel.setMinimumHeight(170)
        self.hist_panel.setMaximumHeight(220)
        dv.addWidget(self.hist_panel)
        av.addWidget(dist_box)

        yield_box = QGroupBox('Pass / Fail Summary')
        ov = QVBoxLayout(yield_box); ov.setContentsMargins(8, 8, 8, 8); ov.setSpacing(6)
        self.yield_donut_panel = YieldDonutPanel()
        ov.addWidget(self.yield_donut_panel)
        av.addWidget(yield_box)

        mini_heatmap_box = QGroupBox('Mini Heatmap')
        mhv = QVBoxLayout(mini_heatmap_box); mhv.setContentsMargins(8, 8, 8, 8); mhv.setSpacing(6)
        mini_summary = QLabel('Compact wafer overview using pass / near-limit / fail colors.')
        mini_summary.setWordWrap(True)
        mini_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        mhv.addWidget(mini_summary)
        self.mini_heatmap_panel = MiniHeatmapPanel()
        mhv.addWidget(self.mini_heatmap_panel)
        av.addWidget(mini_heatmap_box, stretch=1)
        tabs.addTab(self.analytics_panel, 'Analysis')
        tabs.addTab(self.stats_panel, 'Statistics')
        tabs.addTab(self.detail_panel, 'Die Detail')
        mh.addWidget(right)
        self.main_tabs.addTab(wafer_page, 'Wafer View')

        self.raw_tab = QWidget()
        raw_root = QHBoxLayout(self.raw_tab); raw_root.setSpacing(12); raw_root.setContentsMargins(8, 8, 8, 8)

        raw_left = QWidget(); raw_left.setFixedWidth(270)
        raw_left.setStyleSheet(f'background:{T["bg_panel"]};border:1px solid {T["border"]};border-radius:10px;')
        rlv = QVBoxLayout(raw_left); rlv.setSpacing(8); rlv.setContentsMargins(10, 10, 10, 10)
        raw_title = QLabel('Raw File Selection')
        raw_title.setStyleSheet(f'font-weight:700;color:{T["accent_dark"]};')
        rlv.addWidget(raw_title)
        raw_box = QWidget()
        raw_box.setStyleSheet('background:transparent;border:none;')
        rsv = QVBoxLayout(raw_box); rsv.setContentsMargins(0, 0, 0, 0); rsv.setSpacing(6)
        self.raw_data_summary = QLabel('Load a wafer file or batch folder to inspect raw KDF contents.')
        self.raw_data_summary.setWordWrap(True)
        self.raw_data_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        rsv.addWidget(self.raw_data_summary)
        self.raw_current_radio = QRadioButton('Current wafer file')
        self.raw_current_radio.toggled.connect(self._on_raw_selection_changed)
        rsv.addWidget(self.raw_current_radio)
        self.raw_batch_scroll = QScrollArea()
        self.raw_batch_scroll.setWidgetResizable(True)
        self.raw_batch_scroll.setFrameShape(QScrollArea.NoFrame)
        self.raw_batch_wrap = QWidget()
        self.raw_batch_layout = QVBoxLayout(self.raw_batch_wrap)
        self.raw_batch_layout.setContentsMargins(0, 0, 0, 0)
        self.raw_batch_layout.setSpacing(4)
        self.raw_batch_layout.addStretch()
        self.raw_batch_scroll.setWidget(self.raw_batch_wrap)
        rsv.addWidget(self.raw_batch_scroll, stretch=1)
        self.raw_button_group = QButtonGroup(self)
        self.raw_button_group.setExclusive(True)
        rlv.addWidget(raw_box, stretch=1)
        raw_root.addWidget(raw_left)

        raw_view_wrap = QWidget()
        raw_view_wrap.setStyleSheet(
            f'background:{T["bg_panel"]};border:1px solid {T["border"]};border-radius:12px;'
        )
        rvw = QVBoxLayout(raw_view_wrap); rvw.setContentsMargins(8, 8, 8, 8); rvw.setSpacing(6)
        self.raw_data_title = QLabel('Raw table view')
        self.raw_data_title.setStyleSheet(f'font-weight:700;color:{T["accent_dark"]};')
        rvw.addWidget(self.raw_data_title)
        self.raw_data_view = QTableWidget(0, 0)
        self.raw_data_view.setEditTriggers(QTableWidget.NoEditTriggers)
        self.raw_data_view.setAlternatingRowColors(True)
        self.raw_data_view.setSelectionBehavior(QTableWidget.SelectRows)
        self.raw_data_view.verticalHeader().setVisible(False)
        self.raw_data_view.setSortingEnabled(False)
        rvw.addWidget(self.raw_data_view, stretch=1)
        raw_root.addWidget(raw_view_wrap, stretch=1)
        self.main_tabs.addTab(self.raw_tab, 'Raw Data')

        self.batch_tab = QWidget()
        batch_outer = QVBoxLayout(self.batch_tab)
        batch_outer.setContentsMargins(0, 0, 0, 0)
        batch_outer.setSpacing(0)
        self.batch_scroll = QScrollArea()
        self.batch_scroll.setWidgetResizable(True)
        self.batch_scroll.setFrameShape(QScrollArea.NoFrame)
        batch_outer.addWidget(self.batch_scroll)
        batch_content = QWidget()
        self.batch_scroll.setWidget(batch_content)
        btv = QVBoxLayout(batch_content); btv.setContentsMargins(8, 8, 8, 8); btv.setSpacing(8)
        self.batch_sections = QTabWidget()
        btv.addWidget(self.batch_sections)
        batch_analytics_page = QWidget()
        bav = QVBoxLayout(batch_analytics_page); bav.setContentsMargins(8, 8, 8, 8); bav.setSpacing(8)
        self.batch_sections.addTab(batch_analytics_page, 'Analytics')
        batch_limits_page = QWidget()
        blv = QVBoxLayout(batch_limits_page); blv.setContentsMargins(8, 8, 8, 8); blv.setSpacing(8)
        self.batch_sections.addTab(batch_limits_page, 'Limits')

        controls = QGroupBox('Batch Controls')
        ch = QHBoxLayout(controls); ch.setSpacing(8)
        self.batch_load_btn = QPushButton('Load Batch Folder…')
        self.batch_load_btn.setObjectName('primary')
        self.batch_load_btn.clicked.connect(self.open_batch_folder)
        ch.addWidget(self.batch_load_btn)
        self.batch_mkey_combo = ArrowComboBox()
        self.batch_mkey_combo.setMinimumHeight(34)
        self.batch_mkey_combo.currentTextChanged.connect(self._update_batch_table)
        ch.addWidget(self.batch_mkey_combo, stretch=1)
        self.batch_design_mode_combo = ArrowComboBox()
        self.batch_design_mode_combo.setMinimumHeight(34)
        self.batch_design_mode_combo.addItem('All designs (aggregate)', None)
        self.batch_design_mode_combo.currentIndexChanged.connect(self._update_batch_table)
        ch.addWidget(self.batch_design_mode_combo)
        self.batch_sort_combo = ArrowComboBox()
        self.batch_sort_combo.setMinimumHeight(34)
        self.batch_sort_combo.addItems(['Yield (high to low)', 'Yield (low to high)', 'Wafer name'])
        self.batch_sort_combo.currentTextChanged.connect(self._update_batch_table)
        ch.addWidget(self.batch_sort_combo)
        self.batch_compare_btn = QPushButton('Compare Selected')
        self.batch_compare_btn.clicked.connect(self._compare_selected_wafers)
        ch.addWidget(self.batch_compare_btn)
        self.batch_export_report_btn = QPushButton('Export Batch Report…')
        self.batch_export_report_btn.clicked.connect(self._export_batch_report)
        ch.addWidget(self.batch_export_report_btn)
        bav.addWidget(controls)

        self.batch_progress = QProgressBar()
        self.batch_progress.setMinimum(0)
        self.batch_progress.setValue(0)
        self.batch_progress.setVisible(False)
        bav.addWidget(self.batch_progress)

        self.batch_summary = QLabel('Open a folder with KDF files to analyze a batch.')
        self.batch_summary.setWordWrap(True)
        self.batch_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        bav.addWidget(self.batch_summary)

        self.batch_table = QTableWidget(0, 12)
        self.batch_table.setHorizontalHeaderLabels([
            'Wafer File', 'Lot', 'Designs', 'Design Used', 'Sites', 'Mean', 'Median', 'Std Dev',
            'Pass', 'Fail', 'Yield', 'Status'
        ])
        bth = self.batch_table.horizontalHeader()
        bth.setSectionResizeMode(0, QHeaderView.Stretch)
        for col in (1, 2, 3, 4, 8, 9, 10, 11):
            bth.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        self.batch_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.batch_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.batch_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.batch_table.verticalHeader().setVisible(False)
        self.batch_table.setShowGrid(False)
        self.batch_table.setMinimumHeight(180)
        self.batch_table.setMaximumHeight(260)
        self.batch_table.setSortingEnabled(False)
        self.batch_table.horizontalHeader().setSectionsClickable(False)
        self.batch_table.horizontalHeader().setSortIndicatorShown(False)
        self.batch_table.itemDoubleClicked.connect(self._open_batch_selected_wafer)
        self.batch_table.itemSelectionChanged.connect(self._compare_selected_wafers)
        bav.addWidget(self.batch_table)

        radial_box = QGroupBox('Within-Wafer Radial Analysis')
        rvb = QVBoxLayout(radial_box); rvb.setContentsMargins(8, 8, 8, 8)
        self.batch_radial_summary = QLabel('Center vs edge ring analysis for process non-uniformity.')
        self.batch_radial_summary.setWordWrap(True)
        self.batch_radial_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        rvb.addWidget(self.batch_radial_summary)
        self.batch_radial_table = QTableWidget(0, 5)
        self.batch_radial_table.setHorizontalHeaderLabels(['Wafer', 'Center Mean', 'Edge Mean', 'Edge-Center', 'Edge Fail %'])
        self.batch_radial_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.batch_radial_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.batch_radial_table.verticalHeader().setVisible(False)
        self.batch_radial_table.setMaximumHeight(220)
        self.batch_radial_table.setSortingEnabled(False)
        self.batch_radial_table.horizontalHeader().setSectionsClickable(False)
        self.batch_radial_table.horizontalHeader().setSortIndicatorShown(False)
        rvb.addWidget(self.batch_radial_table)
        bav.addWidget(radial_box)

        golden_box = QGroupBox('Golden Wafer Scoring')
        gv = QVBoxLayout(golden_box); gv.setContentsMargins(8, 8, 8, 8)
        gh = QHBoxLayout()
        self.batch_golden_combo = ArrowComboBox()
        self.batch_golden_combo.currentTextChanged.connect(self._update_golden_table)
        gh.addWidget(QLabel('Golden wafer:'))
        gh.addWidget(self.batch_golden_combo, stretch=1)
        gv.addLayout(gh)
        self.batch_golden_summary = QLabel('Score each wafer against a golden reference profile.')
        self.batch_golden_summary.setWordWrap(True)
        self.batch_golden_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        gv.addWidget(self.batch_golden_summary)
        self.batch_golden_table = QTableWidget(0, 4)
        self.batch_golden_table.setHorizontalHeaderLabels(['Wafer', 'Match Score', 'Common Dies', 'Avg |Δ|'])
        self.batch_golden_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.batch_golden_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.batch_golden_table.verticalHeader().setVisible(False)
        self.batch_golden_table.setMaximumHeight(220)
        self.batch_golden_table.setSortingEnabled(False)
        self.batch_golden_table.horizontalHeader().setSectionsClickable(False)
        self.batch_golden_table.horizontalHeader().setSortIndicatorShown(False)
        gv.addWidget(self.batch_golden_table)
        bav.addWidget(golden_box)

        self.batch_compare_summary = QLabel('Select two or more wafers to compare.')
        self.batch_compare_summary.setWordWrap(True)
        self.batch_compare_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        bav.addWidget(self.batch_compare_summary)

        compare_maps_wrap = QWidget()
        compare_maps_row = QHBoxLayout(compare_maps_wrap)
        compare_maps_row.setContentsMargins(0, 0, 0, 0)
        compare_maps_row.setSpacing(8)
        self.batch_compare_cards = []
        for i in range(3):
            card = QGroupBox(f'Wafer {i + 1}')
            cv = QVBoxLayout(card); cv.setContentsMargins(8, 8, 8, 8); cv.setSpacing(6)
            title = QLabel('Not selected')
            title.setStyleSheet(f'font-weight:700;color:{T["accent_dark"]};')
            meta = QLabel('Select rows in the table to compare.')
            meta.setWordWrap(True)
            meta.setStyleSheet(f'font-size:11px;color:{T["text_secondary"]};')
            canvas = WaferCanvas()
            canvas.setMinimumSize(300, 300)
            cv.addWidget(title)
            cv.addWidget(meta)
            cv.addWidget(canvas, stretch=1)
            compare_maps_row.addWidget(card, stretch=1)
            self.batch_compare_cards.append({
                'card': card,
                'title': title,
                'meta': meta,
                'canvas': canvas,
            })
        bav.addWidget(compare_maps_wrap)
        lot_trend_box = QGroupBox('Lot Trend (Yield)')
        ltv = QVBoxLayout(lot_trend_box); ltv.setContentsMargins(8, 8, 8, 8)
        self.batch_trend_panel = YieldTrendPanel()
        ltv.addWidget(self.batch_trend_panel)
        bav.addWidget(lot_trend_box)

        fail_site_box = QGroupBox('Common Fail Site Finder')
        fsv = QVBoxLayout(fail_site_box); fsv.setContentsMargins(8, 8, 8, 8)
        self.batch_fail_site_summary = QLabel('Fail frequency heatmap across wafers.')
        self.batch_fail_site_summary.setWordWrap(True)
        self.batch_fail_site_summary.setStyleSheet(f'color:{T["text_secondary"]};font-size:12px;')
        fsv.addWidget(self.batch_fail_site_summary)
        self.batch_fail_site_panel = BatchFailHeatmapPanel()
        fsv.addWidget(self.batch_fail_site_panel)
        bav.addWidget(fail_site_box)
        bav.addStretch()

        limits_box = QGroupBox('Batch Measurement Limits')
        ll = QVBoxLayout(limits_box); ll.setSpacing(8)
        row1 = QHBoxLayout()
        row1.addWidget(QLabel('Spec Low'))
        self.batch_low_edit = QLineEdit(); self.batch_low_edit.setPlaceholderText('no limit')
        row1.addWidget(self.batch_low_edit)
        row1.addWidget(QLabel('Spec High'))
        self.batch_high_edit = QLineEdit(); self.batch_high_edit.setPlaceholderText('no limit')
        row1.addWidget(self.batch_high_edit)
        ll.addLayout(row1)

        self.batch_prod_toggle = QCheckBox('Use production limits')
        self.batch_prod_toggle.toggled.connect(self._on_batch_prod_toggle)
        ll.addWidget(self.batch_prod_toggle)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel('Prod Low'))
        self.batch_prod_low_edit = QLineEdit(); self.batch_prod_low_edit.setPlaceholderText('no limit')
        row2.addWidget(self.batch_prod_low_edit)
        row2.addWidget(QLabel('Prod High'))
        self.batch_prod_high_edit = QLineEdit(); self.batch_prod_high_edit.setPlaceholderText('no limit')
        row2.addWidget(self.batch_prod_high_edit)
        ll.addLayout(row2)

        br = QHBoxLayout()
        self.batch_limits_apply_btn = QPushButton('Apply Limits')
        self.batch_limits_apply_btn.setObjectName('primary')
        self.batch_limits_apply_btn.clicked.connect(self._apply_batch_limits)
        self.batch_limits_clear_btn = QPushButton('Clear Limits')
        self.batch_limits_clear_btn.clicked.connect(self._clear_batch_limits)
        br.addWidget(self.batch_limits_apply_btn)
        br.addWidget(self.batch_limits_clear_btn)
        ll.addLayout(br)
        blv.addWidget(limits_box)
        blv.addStretch()

        self.main_tabs.addTab(self.batch_tab, 'Batch Analysis')

        self.status = QStatusBar(); self.setStatusBar(self.status)
        self.status.showMessage('  Ready  ·  open a KDF file to begin')

    # ── file loading ──────────────────────────────────────────────────────────

    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 'Open KDF File', '',
            'KDF Files (*.kdf *.KDF);;All Files (*)')
        if path:
            self._load_kdf(path)

    def open_batch_folder(self):
        folder = QFileDialog.getExistingDirectory(self, 'Open Batch Folder', '')
        if not folder:
            return
        self._load_batch_folder(folder)

    def reset_all(self):
        ans = QMessageBox.question(
            self,
            'Reset Everything',
            'Are you sure you want to clear all loaded wafer and batch data?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if ans != QMessageBox.Yes:
            return

        self._header = {}
        self._sites = []
        self._mkeys = []
        self._limits = {}
        self._current_mkey = None
        self._filepath = None
        self._all_subsites = []
        self._current_sub = None
        self._batch_records = []
        self._batch_rows = []
        self._batch_dir = None
        self._use_prod_limits = False

        self.lbl_file.setText('No file loaded  ')
        self.lbl_lot.setText('—')
        self.lbl_sys.setText('—')
        self.lbl_stt.setText('—')
        self.lbl_cnt.setText('—')
        self.lbl_tests.setText('—')
        self.lbl_dsn.setText('—')
        self.setWindowTitle('Wafer Map Viewer')

        self.design_combo.blockSignals(True)
        self.design_combo.clear()
        self.design_combo.blockSignals(False)
        self.mkey_combo.blockSignals(True)
        self.mkey_combo.clear()
        self.mkey_combo.blockSignals(False)
        self.low_edit.clear()
        self.high_edit.clear()
        self.prod_toggle.setChecked(False)
        self.prod_low_edit.clear()
        self.prod_high_edit.clear()

        self.batch_mkey_combo.blockSignals(True)
        self.batch_mkey_combo.clear()
        self.batch_mkey_combo.blockSignals(False)
        self.batch_design_mode_combo.blockSignals(True)
        self.batch_design_mode_combo.clear()
        self.batch_design_mode_combo.addItem('All designs (aggregate)', None)
        self.batch_design_mode_combo.blockSignals(False)
        self.batch_golden_combo.blockSignals(True)
        self.batch_golden_combo.clear()
        self.batch_golden_combo.blockSignals(False)
        self.batch_low_edit.clear()
        self.batch_high_edit.clear()
        self.batch_prod_toggle.setChecked(False)
        self.batch_prod_low_edit.clear()
        self.batch_prod_high_edit.clear()
        self.batch_table.setRowCount(0)
        self.batch_radial_table.setRowCount(0)
        self.batch_golden_table.setRowCount(0)
        self.batch_compare_summary.setText('Select two or more wafers to compare.')
        self.batch_summary.setText('Open a folder with KDF files to analyze a batch.')
        self.batch_radial_summary.setText('Center vs edge ring analysis for process non-uniformity.')
        self.batch_golden_summary.setText('Score each wafer against a golden reference profile.')
        self._clear_compare_cards()
        self.canvas.load([], {}, None, None, mkey='')
        self.hist_panel.set_data([], None, None)
        if hasattr(self, 'yield_donut_panel'):
            self.yield_donut_panel.set_data(0, 0, 0)
        if hasattr(self, 'mini_heatmap_panel'):
            self.mini_heatmap_panel.set_data([], None, None)
        self.cpk_label.setText('Cp/Cpk: N/A')
        self.batch_trend_panel.set_data([])
        self.batch_fail_site_panel.set_data([])
        self.batch_fail_site_summary.setText('Fail frequency heatmap across wafers.')
        self._set_raw_data_path(None)
        self._populate_raw_selector()

        self._update_ui_state()
        self.status.showMessage('  Reset complete')

    def _set_active_batch_tab(self):
        if hasattr(self, 'main_tabs'):
            idx = self.main_tabs.indexOf(self.batch_tab)
            if idx >= 0:
                self.main_tabs.setCurrentIndex(idx)

    def _set_active_wafer_tab(self):
        if hasattr(self, 'main_tabs'):
            self.main_tabs.setCurrentIndex(0)

    def _clear_raw_batch_buttons(self):
        if not hasattr(self, 'raw_batch_layout'):
            return
        while self.raw_batch_layout.count():
            item = self.raw_batch_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                self.raw_button_group.removeButton(widget)
                widget.deleteLater()
        self.raw_batch_layout.addStretch()

    def _populate_raw_selector(self):
        if not hasattr(self, 'raw_current_radio'):
            return
        self.raw_current_radio.blockSignals(True)
        self.raw_current_radio.setChecked(False)
        self.raw_current_radio.setEnabled(bool(self._filepath and os.path.isfile(self._filepath)))
        label = 'Current wafer file'
        if self._filepath:
            label = f'Current wafer file  ·  {os.path.basename(self._filepath)}'
        self.raw_current_radio.setText(label)
        self.raw_current_radio.blockSignals(False)

        self._clear_raw_batch_buttons()
        batch_paths = []
        for rec in self._batch_records:
            path = rec.get('path')
            if path and os.path.isfile(path):
                batch_paths.append(path)
        for path in batch_paths:
            btn = QRadioButton(os.path.basename(path))
            btn.setProperty('raw_path', path)
            btn.toggled.connect(self._on_raw_selection_changed)
            self.raw_button_group.addButton(btn)
            self.raw_batch_layout.insertWidget(self.raw_batch_layout.count() - 1, btn)

        if self._filepath and os.path.isfile(self._filepath):
            self.raw_current_radio.setChecked(True)
            self._set_raw_data_path(self._filepath)
        elif batch_paths:
            first = self.raw_button_group.buttons()[0]
            first.setChecked(True)
            self._set_raw_data_path(batch_paths[0])
        else:
            self._set_raw_data_path(None)

    def _set_raw_data_path(self, path: str | None):
        self._raw_data_path = path
        if not hasattr(self, 'raw_data_view'):
            return
        if not path:
            self.raw_data_title.setText('Raw table view')
            self.raw_data_summary.setText('Load a wafer file to view one row per site/subsite.')
            self.raw_data_view.setRowCount(0)
            self.raw_data_view.setColumnCount(0)
            return
        try:
            header, sites, mkeys, _tests = parse_kdf(path)
        except Exception as e:
            self.raw_data_title.setText('Raw table view')
            self.raw_data_summary.setText(f'Could not parse raw file: {e}')
            self.raw_data_view.setRowCount(0)
            self.raw_data_view.setColumnCount(0)
            return
        self._populate_raw_data_table(path, header, sites, mkeys)

    def _populate_raw_data_table(self, path: str, header: dict, sites: list, mkeys: list):
        columns = ['Site', 'X', 'Y', 'Subsite'] + list(mkeys)
        rows: list[dict] = []
        for site in sites:
            for subsite in sorted(site.get('subsites', {}).keys()):
                row = {
                    'Site': site.get('name', ''),
                    'X': str(site.get('x', '')),
                    'Y': str(site.get('y', '')),
                    'Subsite': str(subsite),
                }
                sub_vals = site.get('subsites', {}).get(subsite, {})
                for mkey in mkeys:
                    val = sub_vals.get(mkey)
                    row[mkey] = '' if val is None else f'{val:.12g}'
                rows.append(row)

        self.raw_data_view.setSortingEnabled(False)
        self.raw_data_view.clear()
        self.raw_data_view.setRowCount(len(rows))
        self.raw_data_view.setColumnCount(len(columns))
        self.raw_data_view.setHorizontalHeaderLabels(columns)
        self.raw_data_view.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.raw_data_view.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.raw_data_view.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.raw_data_view.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        for col in range(4, len(columns)):
            self.raw_data_view.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)

        for r, row in enumerate(rows):
            for c, col_name in enumerate(columns):
                item = QTableWidgetItem(row.get(col_name, ''))
                if c >= 1:
                    item.setTextAlignment(Qt.AlignCenter)
                self.raw_data_view.setItem(r, c, item)

        lot = header.get('LOT', '—')
        self.raw_data_title.setText(f'Raw table view  ·  {os.path.basename(path)}')
        self.raw_data_summary.setText(
            f'One row per site/subsite  ·  lot {lot}  ·  rows: {len(rows)}  ·  metrics: {len(mkeys)}'
        )

    def _on_raw_selection_changed(self, checked: bool):
        if not checked:
            return
        sender = self.sender()
        if sender is self.raw_current_radio:
            self._set_raw_data_path(self._filepath if self._filepath and os.path.isfile(self._filepath) else None)
            return
        if isinstance(sender, QRadioButton):
            self._set_raw_data_path(sender.property('raw_path'))

    def _sync_batch_limit_controls(self):
        if not hasattr(self, 'batch_low_edit'):
            return
        mkey = self.batch_mkey_combo.currentText().strip()
        lo, hi, prod_lo, prod_hi = self._limits.get(mkey, (None, None, None, None))

        # Preserve whatever the user currently has in the batch limit fields.
        # Switching measurement should not "swap" the limits to the newly
        # selected measurement's stored values (often empty).
        ui_has_any = any(
            (self.batch_low_edit.text().strip(),
             self.batch_high_edit.text().strip(),
             self.batch_prod_low_edit.text().strip(),
             self.batch_prod_high_edit.text().strip())
        )
        if not ui_has_any:
            self.batch_low_edit.setText('' if lo is None else str(lo))
            self.batch_high_edit.setText('' if hi is None else str(hi))
            self.batch_prod_low_edit.setText('' if prod_lo is None else str(prod_lo))
            self.batch_prod_high_edit.setText('' if prod_hi is None else str(prod_hi))
        self.batch_prod_toggle.blockSignals(True)
        self.batch_prod_toggle.setChecked(self._use_prod_limits)
        self.batch_prod_toggle.blockSignals(False)
        self.batch_prod_low_edit.setEnabled(bool(self._use_prod_limits))
        self.batch_prod_high_edit.setEnabled(bool(self._use_prod_limits))

    def _on_continuous_heatmap_toggled(self, checked: bool):
        self.canvas.set_continuous_mode(bool(checked))
        self._update_wafer_analytics()

    def _update_wafer_analytics(self, *_args):
        if not hasattr(self, 'hist_panel'):
            return
        if not self._sites or not self._current_mkey:
            self.hist_panel.set_data([], None, None)
            if hasattr(self, 'yield_donut_panel'):
                self.yield_donut_panel.set_data(0, 0, 0)
            if hasattr(self, 'mini_heatmap_panel'):
                self.mini_heatmap_panel.set_data([], None, None)
            self.cpk_label.setText('Cp/Cpk: N/A')
            return
        lo, hi, _prod_lo, _prod_hi = self._limits.get(self._current_mkey, (None, None, None, None))
        finite_vals = []
        passed = 0
        failed = 0
        warned = 0
        mini_points = []
        for s in self._sites:
            v = get_site_value(s, self._current_mkey, self._current_sub)
            if v is None or not math.isfinite(v):
                mini_points.append({'x': s.get('x', 0), 'y': s.get('y', 0), 'value': None})
                continue
            finite_vals.append(v)
            is_fail = (lo is not None and v < lo) or (hi is not None and v > hi)
            if is_fail:
                failed += 1
                mini_points.append({'x': s.get('x', 0), 'y': s.get('y', 0), 'value': v})
                continue
            passed += 1
            near_lo = lo is not None and abs(v - lo) <= max(1e-18, abs(lo) * 0.02)
            near_hi = hi is not None and abs(v - hi) <= max(1e-18, abs(hi) * 0.02)
            if near_lo or near_hi:
                warned += 1
            mini_points.append({'x': s.get('x', 0), 'y': s.get('y', 0), 'value': v})
        self.hist_panel.set_data(finite_vals, lo, hi)
        if hasattr(self, 'yield_donut_panel'):
            self.yield_donut_panel.set_data(passed, failed, warned)
        if hasattr(self, 'mini_heatmap_panel'):
            vmin = min(finite_vals) if finite_vals else None
            vmax = max(finite_vals) if finite_vals else None
            self.mini_heatmap_panel.set_data(mini_points, vmin, vmax)

        if finite_vals and lo is not None and hi is not None:
            mean_v = statistics.mean(finite_vals)
            std_v = statistics.pstdev(finite_vals)
            if std_v > 0:
                cp = (hi - lo) / (6.0 * std_v)
                cpk = min((hi - mean_v) / (3.0 * std_v), (mean_v - lo) / (3.0 * std_v))
                self.cpk_label.setText(f'Cp: {cp:.3f}  ·  Cpk: {cpk:.3f}  ·  Mean: {si_fmt(mean_v)}  ·  Std: {si_fmt(std_v)}')
            else:
                self.cpk_label.setText('Cp/Cpk: undefined (std dev is zero)')
        else:
            self.cpk_label.setText('Cp/Cpk: set both Low/High limits and valid data to compute')

    def _update_batch_trend(self, rows: list[dict]):
        pts = []
        ordered_rows = sorted(rows, key=lambda row: row.get('rec', {}).get('load_index', 0))
        for idx, row in enumerate(ordered_rows, start=1):
            y = row.get('yield_num', -1.0)
            if y is None or y < 0:
                continue
            pts.append((f'Wafer {idx}', float(y)))
        self.batch_trend_panel.set_data(pts)

    def _update_common_fail_site_map(self, rows: list[dict], mkey: str, lo, hi, use_prod: bool, prod_lo, prod_hi):
        agg = defaultdict(lambda: {'x': 0, 'y': 0, 'fail': 0, 'total': 0})
        for row in rows:
            rec = row['rec']
            d_text = str(row.get('design_used', 'All'))
            design = None if d_text in ('—', 'All') else int(d_text)
            for s in rec.get('sites', []):
                v = get_site_value(s, mkey, design)
                if v is None or not math.isfinite(v):
                    continue
                key = (s['x'], s['y'])
                slot = agg[key]
                slot['x'] = s['x']
                slot['y'] = s['y']
                slot['total'] += 1
                in_spec = (lo is None or v >= lo) and (hi is None or v <= hi)
                in_prod = (prod_lo is None or v >= prod_lo) and (prod_hi is None or v <= prod_hi)
                if not (in_spec and (in_prod if use_prod else True)):
                    slot['fail'] += 1
        points = []
        for slot in agg.values():
            total = slot['total']
            if total <= 0:
                continue
            fail = slot['fail']
            points.append({
                'x': slot['x'],
                'y': slot['y'],
                'fail': fail,
                'total': total,
                'fail_frac': fail / total,
            })
        self.batch_fail_site_panel.set_data(points)
        if points:
            hotspots = sum(1 for p in points if p['fail_frac'] >= 0.5)
            self.batch_fail_site_summary.setText(
                f'Fail frequency map over {len(points)} die coordinates  ·  hotspots (>=50% fail): {hotspots}'
            )
        else:
            self.batch_fail_site_summary.setText('No common fail-site data available for current selection.')

    def _load_batch_folder(self, folder: str):
        try:
            names = sorted(os.listdir(folder))
        except OSError as e:
            QMessageBox.critical(self, 'Folder Error', f'Could not read folder:\n{e}')
            return

        kdf_paths = [
            os.path.join(folder, nm) for nm in names
            if os.path.isfile(os.path.join(folder, nm)) and nm.lower().endswith('.kdf')
        ]
        if not kdf_paths:
            QMessageBox.information(self, 'No KDF files', 'No .kdf files were found in this folder.')
            return
        self._set_active_batch_tab()

        self._batch_records = []
        self._batch_dir = folder
        self.batch_progress.setVisible(True)
        self.batch_progress.setMaximum(len(kdf_paths))
        self.batch_progress.setValue(0)
        self.batch_summary.setText(f'Loading {len(kdf_paths)} wafer files...')
        QApplication.processEvents()

        failures = []
        for idx, path in enumerate(kdf_paths, start=1):
            try:
                header, sites, mkeys, tests = parse_kdf(path)
                subs = all_subsites(sites)
                self._batch_records.append({
                    'path': path,
                    'name': os.path.basename(path),
                    'load_index': idx,
                    'header': header,
                    'sites': sites,
                    'mkeys': mkeys,
                    'tests': tests,
                    'subs': subs,
                })
            except Exception as e:
                failures.append((os.path.basename(path), str(e)))
            self.batch_progress.setValue(idx)
            self.batch_summary.setText(f'Loading {idx}/{len(kdf_paths)} wafer files...')
            QApplication.processEvents()

        if not self._batch_records:
            self.batch_progress.setVisible(False)
            self._update_ui_state()
            detail = '\n'.join([f'- {nm}: {err}' for nm, err in failures[:6]]) or 'No files could be parsed.'
            QMessageBox.critical(
                self, 'Batch Parse Error',
                f'No KDF files could be parsed from this folder.\n\n{detail}'
            )
            return

        mkeys = sorted({mk for rec in self._batch_records for mk in rec['mkeys']})
        batch_subs = sorted({sn for rec in self._batch_records for sn in rec.get('subs', [])})
        self.batch_mkey_combo.blockSignals(True)
        self.batch_mkey_combo.clear()
        self.batch_mkey_combo.addItems(mkeys)
        self.batch_mkey_combo.blockSignals(False)
        self.batch_design_mode_combo.blockSignals(True)
        self.batch_design_mode_combo.clear()
        self.batch_design_mode_combo.addItem('All designs (aggregate)', None)
        for sn in batch_subs:
            self.batch_design_mode_combo.addItem(f'Design {sn}', sn)
        if batch_subs:
            idx = self.batch_design_mode_combo.findData(batch_subs[0])
            if idx >= 0:
                self.batch_design_mode_combo.setCurrentIndex(idx)
        self.batch_design_mode_combo.blockSignals(False)
        if mkeys:
            if self._current_mkey and self._current_mkey in mkeys:
                self.batch_mkey_combo.setCurrentText(self._current_mkey)
            else:
                self.batch_mkey_combo.setCurrentIndex(0)
        self._sync_batch_limit_controls()

        self.batch_progress.setVisible(False)
        self._update_batch_table()
        self._populate_raw_selector()
        self._update_ui_state()

        ok_n = len(self._batch_records)
        msg = f'Loaded batch: {ok_n}/{len(kdf_paths)} wafer files'
        if failures:
            msg += f'  ·  {len(failures)} failed to parse'
        self.status.showMessage(f'  {msg}')

    def _load_kdf(self, path: str):
        preserve_limits = bool(getattr(self, '_preserve_limits_on_next_load', False))
        self._preserve_limits_on_next_load = False
        try:
            header, sites, params, tests = parse_kdf(path)
        except Exception as e:
            QMessageBox.critical(self, 'Parse Error',
                                 f'Could not read KDF file:\n{e}')
            return

        self._filepath     = path
        self._header       = header
        self._sites        = sites
        self._mkeys        = params
        if not preserve_limits:
            self._limits = {}
        self._current_mkey = None

        subs = all_subsites(sites)
        self._all_subsites = subs
        self._current_sub  = subs[0] if len(subs) == 1 else None

        self.lbl_file.setText(f'  {os.path.basename(path)}  ')
        self.lbl_lot.setText(header.get('LOT', '—'))
        self.lbl_sys.setText(header.get('SYS', '—') or header.get('TST', '—'))
        self.lbl_stt.setText(header.get('STT', '—'))
        self.lbl_cnt.setText(str(len(sites)))
        self.lbl_tests.setText(str(len(tests)))
        self.lbl_dsn.setText(str(len(subs)))

        self.setWindowTitle(
            f'Wafer Map Viewer  ·  {header.get("LOT", os.path.basename(path))}')

        # Build design selector (may be sorted by pass count once limits are active).
        self._rebuild_design_combo()

        self.mkey_combo.blockSignals(True)
        self.mkey_combo.clear()
        self.mkey_combo.addItems(params)
        self.mkey_combo.blockSignals(False)

        if params:
            self._current_mkey = params[0]
            self.mkey_combo.setCurrentText(params[0])
            self._refresh_canvas()
        self._update_wafer_analytics()

        self._update_ui_state()
        self.status.showMessage(
            f'  Loaded {len(sites)} sites  ·  '
            f'{len(params)} measurements  ·  '
            f'{len(tests)} tests  ·  '
            f'{len(subs)} design(s)  ·  {os.path.basename(path)}')
        self._populate_raw_selector()
        self._update_batch_table()
        self._set_active_wafer_tab()
        self._sync_batch_limit_controls()

    # ── design / measurement / limits ─────────────────────────────────────────

    def _rebuild_design_combo(self):
        """Populate design selector, optionally sorted by pass count when limits are active."""
        subs = list(self._all_subsites or [])
        if not subs:
            return

        # Preserve current selection if possible.
        prev_sub = self._current_sub

        mkey = self._current_mkey
        lo = hi = prod_lo = prod_hi = None
        if mkey:
            lo, hi, prod_lo, prod_hi = self._limits.get(mkey, (None, None, None, None))
        limits_active = (lo is not None or hi is not None) and bool(mkey)
        prod_active = self._use_prod_limits and (prod_lo is not None or prod_hi is not None)

        def pass_and_total_for(sub_num: int) -> tuple[int, int]:
            passed = 0
            total = 0
            for s in self._sites:
                v = get_site_value(s, mkey, sub_num)
                if v is None:
                    continue
                total += 1
                in_spec = (lo is None or v >= lo) and (hi is None or v <= hi)
                if not in_spec:
                    continue
                if prod_active:
                    in_prod = (prod_lo is None or v >= prod_lo) and (prod_hi is None or v <= prod_hi)
                    if not in_prod:
                        continue
                    passed += 1
                else:
                    passed += 1
            return passed, total

        ordered = subs
        pass_counts: dict[int, int] = {}
        total_counts: dict[int, int] = {}
        if limits_active:
            for sn in subs:
                p, t = pass_and_total_for(sn)
                pass_counts[sn] = p
                total_counts[sn] = t
            ordered = sorted(subs, key=lambda sn: (pass_counts.get(sn, 0), -sn), reverse=True)

        self.design_combo.blockSignals(True)
        self.design_combo.clear()

        for sn in ordered:
            if limits_active:
                pc = pass_counts.get(sn, 0)
                tc = total_counts.get(sn, 0)
                pct = (pc / tc * 100.0) if tc else 0.0
                self.design_combo.addItem(f'Design {sn}  ·  Pass {pc}/{tc} ({pct:.1f}%)', sn)
            else:
                self.design_combo.addItem(f'Design {sn}', sn)

        # Restore selection if possible; otherwise pick first real design (or only).
        idx = self.design_combo.findData(prev_sub)
        if idx < 0:
            idx = 0
        self.design_combo.setCurrentIndex(idx)
        self._current_sub = self.design_combo.itemData(idx)
        self.design_combo.blockSignals(False)

    def _on_design_changed(self, idx: int):
        if idx < 0:
            return
        self._current_sub = self.design_combo.itemData(idx)
        self._refresh_canvas()
        sub = self._current_sub
        label = f'Design {sub}' if sub is not None else 'Design'
        if self._current_mkey:
            self.status.showMessage(
                f'  {self._current_mkey}  ·  {label}  ·  {len(self._sites)} sites')

    def _on_mkey_changed(self, mkey: str):
        if mkey not in self._mkeys:
            return
        if self._current_mkey and not self._commit_limits_from_main_fields(self._current_mkey):
            return

        self._current_mkey = mkey
        # Keep the currently shown limits; only update the stored values
        # for the newly selected measurement.
        if not self._commit_limits_from_main_fields(self._current_mkey):
            return
        if self.batch_mkey_combo.findText(mkey) >= 0:
            self.batch_mkey_combo.setCurrentText(mkey)
        self._sync_batch_limit_controls()
        self._rebuild_design_combo()
        self._refresh_canvas()
        self._update_batch_table()

    def _apply_limits(self):
        if self._current_mkey and not self._commit_limits_from_main_fields(self._current_mkey):
            return
        self._rebuild_design_combo()
        self._refresh_canvas()
        self._sync_batch_limit_controls()
        self._update_batch_table()

    def _clear_limits(self):
        if self._current_mkey:
            ans = QMessageBox.question(
                self,
                'Clear limits',
                f'Clear spec and production limits for {self._current_mkey}?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if ans != QMessageBox.Yes:
                return
        self.low_edit.clear(); self.high_edit.clear()
        self.prod_low_edit.clear(); self.prod_high_edit.clear()
        if self._current_mkey:
            self._limits[self._current_mkey] = (None, None, None, None)
            if self.batch_mkey_combo.currentText().strip() == self._current_mkey:
                self._limits[self.batch_mkey_combo.currentText().strip()] = (None, None, None, None)
        self._rebuild_design_combo()
        self._refresh_canvas()
        self._sync_batch_limit_controls()
        self._update_batch_table()

    def _on_prod_toggle(self, checked: bool):
        self._use_prod_limits = bool(checked)
        self.prod_limits_wrap.setVisible(self._use_prod_limits)
        self.batch_prod_toggle.blockSignals(True)
        self.batch_prod_toggle.setChecked(self._use_prod_limits)
        self.batch_prod_toggle.blockSignals(False)
        self.batch_prod_low_edit.setEnabled(self._use_prod_limits)
        self.batch_prod_high_edit.setEnabled(self._use_prod_limits)
        self._update_ui_state()
        self._rebuild_design_combo()
        self._refresh_canvas()
        self._update_batch_table()

    def _on_batch_prod_toggle(self, checked: bool):
        self._use_prod_limits = bool(checked)
        self.prod_toggle.blockSignals(True)
        self.prod_toggle.setChecked(self._use_prod_limits)
        self.prod_toggle.blockSignals(False)
        self.prod_limits_wrap.setVisible(self._use_prod_limits)
        self.batch_prod_low_edit.setEnabled(self._use_prod_limits)
        self.batch_prod_high_edit.setEnabled(self._use_prod_limits)
        self._update_ui_state()
        self._rebuild_design_combo()
        self._refresh_canvas()
        self._update_batch_table()

    def _apply_batch_limits(self):
        mkey = self.batch_mkey_combo.currentText().strip() or self._current_mkey
        if not mkey:
            QMessageBox.information(self, 'No measurement', 'Select a measurement first.')
            return
        limits = self._collect_limits(
            self.batch_low_edit.text(), self.batch_high_edit.text(),
            self.batch_prod_low_edit.text(), self.batch_prod_high_edit.text(),
            'Spec Low', 'Spec High',
        )
        if limits is INVALID_LIMIT:
            return
        lo, hi, prod_lo, prod_hi = limits
        self._limits[mkey] = limits
        if self._current_mkey == mkey:
            self.low_edit.setText('' if lo is None else str(lo))
            self.high_edit.setText('' if hi is None else str(hi))
            self.prod_low_edit.setText('' if prod_lo is None else str(prod_lo))
            self.prod_high_edit.setText('' if prod_hi is None else str(prod_hi))
        self._rebuild_design_combo()
        self._refresh_canvas()
        self._update_batch_table()

    def _clear_batch_limits(self):
        mkey = self.batch_mkey_combo.currentText().strip() or self._current_mkey
        if not mkey:
            return
        self.batch_low_edit.clear()
        self.batch_high_edit.clear()
        self.batch_prod_low_edit.clear()
        self.batch_prod_high_edit.clear()
        self._limits[mkey] = (None, None, None, None)
        if self._current_mkey == mkey:
            self.low_edit.clear()
            self.high_edit.clear()
            self.prod_low_edit.clear()
            self.prod_high_edit.clear()
        self._rebuild_design_combo()
        self._refresh_canvas()
        self._update_batch_table()

    def _parse_limit(self, text: str, label: str = ''):
        text = text.strip()
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            if label:
                QMessageBox.warning(self, 'Invalid value',
                                    f'{label} limit must be a number (e.g. 9.5 or 1e-12).')
            return INVALID_LIMIT

    def _collect_limits(self, low_text: str, high_text: str, prod_low_text: str, prod_high_text: str,
                        low_label: str, high_label: str):
        lo = self._parse_limit(low_text, low_label)
        hi = self._parse_limit(high_text, high_label)
        prod_lo = self._parse_limit(prod_low_text, 'Prod Low')
        prod_hi = self._parse_limit(prod_high_text, 'Prod High')
        if INVALID_LIMIT in (lo, hi, prod_lo, prod_hi):
            return INVALID_LIMIT
        return lo, hi, prod_lo, prod_hi

    def _commit_limits_from_main_fields(self, mkey: str) -> bool:
        limits = self._collect_limits(
            self.low_edit.text(), self.high_edit.text(),
            self.prod_low_edit.text(), self.prod_high_edit.text(),
            'Low', 'High',
        )
        if limits is INVALID_LIMIT:
            return False
        self._limits[mkey] = limits
        if self.batch_mkey_combo.currentText().strip() == mkey:
            self._limits[self.batch_mkey_combo.currentText().strip()] = limits
        return True

    def _refresh_canvas(self):
        if not self._sites or not self._current_mkey:
            return
        mkey   = self._current_mkey
        sub    = self._current_sub
        lo, hi, prod_lo, prod_hi = self._limits.get(mkey, (None, None, None, None))

        values = {s['name']: get_site_value(s, mkey, sub) for s in self._sites}

        self.canvas.load(
            self._sites, values, lo, hi, mkey=mkey,
            prod_lo=prod_lo, prod_hi=prod_hi, show_prod=self._use_prod_limits
        )
        self.stats_panel.update_stats(values, lo, hi)
        self._update_wafer_analytics()

        sub_label = f'design {sub}' if sub is not None else 'design'
        lo_s = '—' if lo is None else str(lo)
        hi_s = '—' if hi is None else str(hi)
        plo_s = '—' if prod_lo is None else str(prod_lo)
        phi_s = '—' if prod_hi is None else str(prod_hi)
        prod_label = 'off'
        if self._use_prod_limits:
            prod_label = f'Low = {plo_s}  High = {phi_s}'
        self.status.showMessage(
            f'  {mkey}  ·  {sub_label}'
            f'  ·  Low = {lo_s}  High = {hi_s}'
            f'  ·  Prod: {prod_label}'
            f'  ·  {len(self._sites)} sites')

    def _on_die_clicked(self, site: dict):
        self.detail_panel.show_site(site)

    # ── batch analysis ────────────────────────────────────────────────────────

    def _batch_design_for_record(self, rec: dict) -> int | None:
        subs = rec.get('subs') or []
        mode = self.batch_design_mode_combo.currentData() if hasattr(self, 'batch_design_mode_combo') else None
        if mode is None:
            return None  # aggregate across all designs
        try:
            mode_int = int(mode)
        except (TypeError, ValueError):
            return None
        if mode_int in subs:
            return mode_int
        # If a specific design is not present on this wafer, fall back to aggregate.
        return None

    def _batch_values_for_record(self, rec: dict, mkey: str):
        design = self._batch_design_for_record(rec)
        values = {}
        for s in rec.get('sites', []):
            values[s['name']] = get_site_value(s, mkey, design)
        return design, values

    def _update_batch_table(self, *_args):
        if not hasattr(self, 'batch_table'):
            return
        self._sync_batch_limit_controls()
        prev_selected_paths = set()
        sel_model = self.batch_table.selectionModel()
        if sel_model:
            for idx in sel_model.selectedRows():
                it = self.batch_table.item(idx.row(), 0)
                if it is not None:
                    path = it.data(Qt.UserRole)
                    if path:
                        prev_selected_paths.add(path)
        if not self._batch_records:
            self.batch_table.setRowCount(0)
            self.batch_compare_summary.setText('Select two or more wafers to compare.')
            self._clear_compare_cards()
            self._batch_rows = []
            self.batch_trend_panel.set_data([])
            self.batch_fail_site_panel.set_data([])
            self.batch_fail_site_summary.setText('Fail frequency heatmap across wafers.')
            return

        mkey = self.batch_mkey_combo.currentText().strip()
        if not mkey:
            self.batch_table.setRowCount(0)
            self.batch_summary.setText(
                f'Loaded {len(self._batch_records)} wafers. No measurement selected.'
            )
            self.batch_compare_summary.setText('Select two or more wafers to compare.')
            self._clear_compare_cards()
            self._batch_rows = []
            self.batch_trend_panel.set_data([])
            self.batch_fail_site_panel.set_data([])
            self.batch_fail_site_summary.setText('Select a measurement to compute fail-site heatmap.')
            return

        # Use the currently visible batch limit fields as the effective
        # limits. This keeps the fail-site finder consistent when the
        # measurement dropdown changes.
        limits = self._collect_limits(
            self.batch_low_edit.text(), self.batch_high_edit.text(),
            self.batch_prod_low_edit.text(), self.batch_prod_high_edit.text(),
            'Spec Low', 'Spec High',
        )
        if limits is INVALID_LIMIT:
            return
        lo, hi, prod_lo, prod_hi = limits
        use_prod = self._use_prod_limits and (prod_lo is not None or prod_hi is not None)

        rows = []
        total_valid = total_pass = 0
        for rec in self._batch_records:
            if mkey not in rec.get('mkeys', []):
                rows.append({
                    'rec': rec, 'designs': len(rec.get('subs', [])), 'design_used': '—',
                    'sites': len(rec.get('sites', [])), 'mean': 'N/A', 'median': 'N/A', 'std': 'N/A',
                    'pass': 0, 'fail': 0, 'yield': 'N/A', 'status': 'No measurement',
                    'yield_num': -1.0,
                })
                continue

            design, values = self._batch_values_for_record(rec, mkey)
            vals = [v for v in values.values() if v is not None and math.isfinite(v)]
            if vals:
                mean_v = statistics.mean(vals)
                med_v = statistics.median(vals)
                std_v = statistics.pstdev(vals)
            else:
                mean_v = med_v = std_v = None

            passed = 0
            for v in vals:
                in_spec = (lo is None or v >= lo) and (hi is None or v <= hi)
                if not in_spec:
                    continue
                if use_prod:
                    in_prod = (prod_lo is None or v >= prod_lo) and (prod_hi is None or v <= prod_hi)
                    if not in_prod:
                        continue
                passed += 1
            fail = max(0, len(vals) - passed)
            yld = (passed / len(vals) * 100.0) if vals else None
            total_valid += len(vals)
            total_pass += passed

            status = 'OK'
            if len(rec.get('subs', [])) > 1:
                status = 'Multi-design wafer'
            if not vals:
                status = 'No data'

            # radial split: center vs edge
            center_vals = []
            edge_vals = []
            edge_total = 0
            edge_fail = 0
            xs = [s['x'] for s in rec.get('sites', [])]
            ys = [s['y'] for s in rec.get('sites', [])]
            cx = statistics.mean(xs) if xs else 0.0
            cy = statistics.mean(ys) if ys else 0.0
            max_r = max((math.hypot(s['x'] - cx, s['y'] - cy) for s in rec.get('sites', [])), default=1.0)
            for s in rec.get('sites', []):
                v = values.get(s['name'])
                if v is None or not math.isfinite(v):
                    continue
                r = math.hypot(s['x'] - cx, s['y'] - cy) / max_r if max_r > 0 else 0.0
                if r <= 0.5:
                    center_vals.append(v)
                else:
                    edge_vals.append(v)
                    edge_total += 1
                    in_spec = (lo is None or v >= lo) and (hi is None or v <= hi)
                    in_prod = (prod_lo is None or v >= prod_lo) and (prod_hi is None or v <= prod_hi)
                    passed_v = in_spec and (in_prod if use_prod else True)
                    if not passed_v:
                        edge_fail += 1

            center_mean = statistics.mean(center_vals) if center_vals else None
            edge_mean = statistics.mean(edge_vals) if edge_vals else None
            edge_fail_pct = (edge_fail / edge_total * 100.0) if edge_total else None

            rows.append({
                'rec': rec,
                'designs': len(rec.get('subs', [])),
                'design_used': ('All' if self.batch_design_mode_combo.currentData() is None else ('—' if design is None else str(design))),
                'sites': len(rec.get('sites', [])),
                'mean': si_fmt(mean_v) if mean_v is not None else 'N/A',
                'median': si_fmt(med_v) if med_v is not None else 'N/A',
                'std': si_fmt(std_v) if std_v is not None else 'N/A',
                'pass': passed,
                'fail': fail,
                'yield': (f'{yld:.1f}%' if yld is not None else 'N/A'),
                'status': status,
                'yield_num': (-1.0 if yld is None else yld),
                'mean_num': mean_v,
                'std_num': std_v,
                'center_mean': center_mean,
                'edge_mean': edge_mean,
                'edge_fail_pct': edge_fail_pct,
            })

        sort_mode = self.batch_sort_combo.currentText() if hasattr(self, 'batch_sort_combo') else 'Yield (high to low)'
        if sort_mode == 'Yield (low to high)':
            rows.sort(key=lambda r: (9999.0 if r['yield_num'] < 0 else r['yield_num']))
        elif sort_mode == 'Wafer name':
            rows.sort(key=lambda r: r['rec']['name'].lower())
        else:
            rows.sort(key=lambda r: r['yield_num'], reverse=True)

        self.batch_table.blockSignals(True)
        self.batch_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            rec = row['rec']
            vals = [
                rec['name'],
                rec['header'].get('LOT', '—'),
                str(row['designs']),
                row['design_used'],
                str(row['sites']),
                row['mean'],
                row['median'],
                row['std'],
                str(row['pass']),
                str(row['fail']),
                row['yield'],
                row['status'],
            ]
            for col, txt in enumerate(vals):
                it = QTableWidgetItem(txt)
                if col in (8, 9):
                    it.setTextAlignment(Qt.AlignCenter)
                if col == 10:
                    it.setTextAlignment(Qt.AlignCenter)
                    if row['yield_num'] >= 90:
                        it.setForeground(QColor(T['pass_fg']))
                    elif row['yield_num'] >= 0 and row['yield_num'] < 70:
                        it.setForeground(QColor(T['fail_fg']))
                    elif row['yield_num'] >= 0:
                        it.setForeground(QColor(T['warn']))
                self.batch_table.setItem(i, col, it)
            # store path in first column item for easy retrieval on double click
            self.batch_table.item(i, 0).setData(Qt.UserRole, rec['path'])
            if rec.get('path') in prev_selected_paths:
                for col in range(self.batch_table.columnCount()):
                    cell = self.batch_table.item(i, col)
                    if cell is not None:
                        cell.setSelected(True)
        self.batch_table.blockSignals(False)

        overall = (total_pass / total_valid * 100.0) if total_valid else 0.0
        prod_txt = 'on' if use_prod else 'off'
        dmode = self.batch_design_mode_combo.currentText() if hasattr(self, 'batch_design_mode_combo') else 'All designs'
        self.batch_summary.setText(
            f'Batch folder: {self._batch_dir or "—"}\n'
            f'Wafers: {len(rows)}  ·  Measurement: {mkey}  ·  Design mode: {dmode}  ·  Prod limits: {prod_txt}\n'
            f'Overall pass: {total_pass}/{total_valid}  ·  Overall yield: {overall:.1f}%\n'
            f'Tip: double-click a wafer row to open it in the wafer map view.'
        )
        self._batch_rows = rows
        self._update_radial_panel(rows)
        self._sync_golden_combo(rows)
        self._update_golden_table()
        self._update_batch_trend(rows)
        self._update_common_fail_site_map(rows, mkey, lo, hi, use_prod, prod_lo, prod_hi)
        if self.batch_table.selectionModel() and self.batch_table.selectionModel().selectedRows():
            self._compare_selected_wafers()

    def _update_radial_panel(self, rows: list[dict]):
        self.batch_radial_table.setRowCount(len(rows))
        edge_higher = 0
        comparable = 0
        for i, row in enumerate(rows):
            c = row['center_mean']
            e = row['edge_mean']
            d = (e - c) if (e is not None and c is not None) else None
            if d is not None:
                comparable += 1
                if d > 0:
                    edge_higher += 1
            vals = [
                row['rec']['name'],
                si_fmt(c) if c is not None else 'N/A',
                si_fmt(e) if e is not None else 'N/A',
                si_fmt(d) if d is not None else 'N/A',
                (f'{row["edge_fail_pct"]:.1f}%' if row['edge_fail_pct'] is not None else 'N/A'),
            ]
            for col, txt in enumerate(vals):
                self.batch_radial_table.setItem(i, col, QTableWidgetItem(txt))
        self.batch_radial_summary.setText(
            f'Radial overview: comparable wafers {comparable}/{len(rows)}  ·  '
            f'Edge>Center in {edge_higher} wafers'
        )

    def _sync_golden_combo(self, rows: list[dict]):
        names = [r['rec']['name'] for r in rows]
        prev = self.batch_golden_combo.currentText()
        self.batch_golden_combo.blockSignals(True)
        self.batch_golden_combo.clear()
        self.batch_golden_combo.addItems(names)
        self.batch_golden_combo.blockSignals(False)
        if prev and prev in names:
            self.batch_golden_combo.setCurrentText(prev)
        elif names:
            self.batch_golden_combo.setCurrentIndex(0)

    def _update_golden_table(self, *_args):
        rows = self._batch_rows or []
        if not rows:
            self.batch_golden_table.setRowCount(0)
            self.batch_golden_summary.setText('Score each wafer against a golden reference profile.')
            return
        gname = self.batch_golden_combo.currentText().strip()
        golden_row = next((r for r in rows if r['rec']['name'] == gname), None)
        if not golden_row:
            self.batch_golden_table.setRowCount(0)
            self.batch_golden_summary.setText('Select a golden wafer to score the batch.')
            return
        mkey = self.batch_mkey_combo.currentText().strip()
        g_design_text = str(golden_row.get('design_used', 'All'))
        g_design = None if g_design_text in ('—', 'All') else int(g_design_text)
        g_values = {
            s['name']: get_site_value(s, mkey, g_design)
            for s in golden_row['rec'].get('sites', [])
        }
        g_sigma_values = [v for v in g_values.values() if v is not None]
        if not g_sigma_values:
            gstd = None
        else:
            gstd = statistics.pstdev(g_sigma_values) or 1.0
        scores = []
        for row in rows:
            d_text = str(row.get('design_used', 'All'))
            design = None if d_text in ('—', 'All') else int(d_text)
            vals = {
                s['name']: get_site_value(s, mkey, design)
                for s in row['rec'].get('sites', [])
            }
            diffs = []
            for name, gv in g_values.items():
                vv = vals.get(name)
                if gv is None or vv is None:
                    continue
                diffs.append(abs(vv - gv))
            common = len(diffs)
            if common:
                avg_abs = statistics.mean(diffs)
                nr = avg_abs / gstd if gstd else 0.0
                score = max(0.0, min(100.0, 100.0 - nr * 25.0))
            else:
                avg_abs = None
                score = 0.0
            scores.append((row['rec']['name'], score, common, avg_abs))
        scores.sort(key=lambda x: x[1], reverse=True)
        self.batch_golden_table.setRowCount(len(scores))
        max_common = 0
        for i, (name, score, common, avg_abs) in enumerate(scores):
            max_common = max(max_common, common)
            vals = [name, f'{score:.1f}', str(common), si_fmt(avg_abs) if avg_abs is not None else 'N/A']
            for col, txt in enumerate(vals):
                it = QTableWidgetItem(txt)
                if col == 1:
                    if score >= 85:
                        it.setForeground(QColor(T['pass_fg']))
                    elif score < 60:
                        it.setForeground(QColor(T['fail_fg']))
                    else:
                        it.setForeground(QColor(T['warn']))
                self.batch_golden_table.setItem(i, col, it)
        if max_common == 0:
            self.batch_golden_summary.setText(
                f'Golden wafer: {gname}  ·  No common die values found for this measurement/design mode.'
            )
        else:
            self.batch_golden_summary.setText(
                f'Golden wafer: {gname}  ·  Scores are based on average absolute delta normalized to golden sigma.'
            )

    def _compare_selected_wafers(self):
        if not hasattr(self, 'batch_table'):
            return
        selected_rows = sorted({idx.row() for idx in self.batch_table.selectionModel().selectedRows()})
        if len(selected_rows) < 2:
            self.batch_compare_summary.setText('Select two or more wafers to compare.')
            self._clear_compare_cards()
            return

        comp = []
        for row in selected_rows:
            name_it = self.batch_table.item(row, 0)
            yld_it = self.batch_table.item(row, 10)
            pass_it = self.batch_table.item(row, 8)
            fail_it = self.batch_table.item(row, 9)
            mean_it = self.batch_table.item(row, 5)
            if not name_it:
                continue
            yld_val = -1.0
            if yld_it and yld_it.text().endswith('%'):
                try:
                    yld_val = float(yld_it.text().rstrip('%'))
                except ValueError:
                    yld_val = -1.0
            row_data = self._batch_rows[row] if (0 <= row < len(self._batch_rows)) else None
            comp.append({
                'name': name_it.text(),
                'yield_num': yld_val,
                'yield_txt': (yld_it.text() if yld_it else 'N/A'),
                'pass_txt': (pass_it.text() if pass_it else '0'),
                'fail_txt': (fail_it.text() if fail_it else '0'),
                'mean_txt': (mean_it.text() if mean_it else 'N/A'),
                'row_data': row_data,
            })

        if len(comp) < 2:
            self.batch_compare_summary.setText('Select two or more wafers to compare.')
            self._clear_compare_cards()
            return

        comp_sorted = sorted(comp, key=lambda x: x['yield_num'], reverse=True)
        best = comp_sorted[0]
        worst = comp_sorted[-1]
        delta = (best['yield_num'] - worst['yield_num']) if best['yield_num'] >= 0 and worst['yield_num'] >= 0 else None
        delta_txt = f'{delta:.1f}%' if delta is not None else 'N/A'
        self.batch_compare_summary.setText(
            f'Comparing {len(comp)} wafers  ·  '
            f'Best yield: {best["name"]} ({best["yield_txt"]})  ·  '
            f'Worst yield: {worst["name"]} ({worst["yield_txt"]})  ·  '
            f'Delta: {delta_txt}'
        )
        self._render_compare_cards(comp[:3])

    def _clear_compare_cards(self):
        if not hasattr(self, 'batch_compare_cards'):
            return
        for idx, card in enumerate(self.batch_compare_cards):
            card['title'].setText(f'Wafer {idx + 1}  ·  Not selected')
            card['meta'].setText('Select rows in the table to compare.')
            card['canvas'].load([], {}, None, None, mkey='')

    def _render_compare_cards(self, comp_rows: list[dict]):
        if not hasattr(self, 'batch_compare_cards'):
            return
        mkey = self.batch_mkey_combo.currentText().strip()
        lo, hi, prod_lo, prod_hi = self._limits.get(mkey, (None, None, None, None))
        use_prod = self._use_prod_limits and (prod_lo is not None or prod_hi is not None)

        for idx, card in enumerate(self.batch_compare_cards):
            if idx >= len(comp_rows):
                card['title'].setText(f'Wafer {idx + 1}  ·  Not selected')
                card['meta'].setText('Select rows in the table to compare.')
                card['canvas'].load([], {}, None, None, mkey='')
                continue

            c = comp_rows[idx]
            row_data = c.get('row_data') or {}
            rec = row_data.get('rec')
            if not rec:
                card['title'].setText(c.get('name', f'Wafer {idx + 1}'))
                card['meta'].setText('No data available.')
                card['canvas'].load([], {}, None, None, mkey='')
                continue

            design = row_data.get('design_used')
            try:
                design_num = int(design) if design not in (None, '—') else None
            except ValueError:
                design_num = None
            values = {s['name']: get_site_value(s, mkey, design_num) for s in rec.get('sites', [])}
            card['canvas'].load(
                rec.get('sites', []), values, lo, hi, mkey=mkey,
                prod_lo=prod_lo, prod_hi=prod_hi, show_prod=use_prod
            )
            card['title'].setText(f'{rec.get("name", "Wafer")}  ·  Yield {c.get("yield_txt", "N/A")}')
            lim_text = f'Spec: {("—" if lo is None else lo)} to {("—" if hi is None else hi)}'
            if use_prod:
                lim_text += f'  ·  Prod: {("—" if prod_lo is None else prod_lo)} to {("—" if prod_hi is None else prod_hi)}'
            card['meta'].setText(
                f'Mean: {c.get("mean_txt", "N/A")}  ·  Pass/Fail: {c.get("pass_txt", "0")}/{c.get("fail_txt", "0")}\n{lim_text}'
            )

    def _open_batch_selected_wafer(self, item: QTableWidgetItem):
        if item is None:
            return
        row = item.row()
        it0 = self.batch_table.item(row, 0)
        if it0 is None:
            return
        path = it0.data(Qt.UserRole)
        if not path or not os.path.isfile(path):
            QMessageBox.warning(self, 'File Missing', 'Selected wafer file is no longer available.')
            return
        # Keep the batch-applied limits when switching to Wafer View.
        self._preserve_limits_on_next_load = True
        self._load_kdf(path)

    def _die_fill_hex(self, v, lo, hi, prod_lo, prod_hi, use_prod):
        # Excel export palette (more saturated / visible than on-screen theme colors).
        PASS = PASS_COLOR
        WARN = WARN_COLOR
        FAIL = FAIL_COLOR
        NEUTRAL = '#546e7a'  # blue-grey
        NODATA = '#7b1fa2'   # purple

        if v is None or not math.isfinite(v):
            return NODATA

        limits_active = (lo is not None or hi is not None)
        if not limits_active:
            return NEUTRAL

        in_spec = (lo is None or v >= lo) and (hi is None or v <= hi)
        if not in_spec:
            return FAIL

        if use_prod:
            in_prod = (prod_lo is None or v >= prod_lo) and (prod_hi is None or v <= prod_hi)
            if not in_prod:
                return WARN

        return PASS

    def export_map_excel(self):
        if not self._sites:
            QMessageBox.information(self, 'Nothing to export', 'Load a KDF file first.')
            return
        if not self._current_mkey:
            QMessageBox.information(self, 'No measurement', 'Select a measurement first.')
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
            from openpyxl.utils import get_column_letter
        except Exception:
            QMessageBox.warning(
                self,
                'Missing dependency',
                'Excel export requires openpyxl.\nInstall with: pip install openpyxl'
            )
            return

        default_name = 'wafer_map.xlsx'
        if self._filepath:
            base = os.path.splitext(os.path.basename(self._filepath))[0]
            default_name = f'{base}_{self._current_mkey.replace("@", "_")}.xlsx'
        out_path, _ = QFileDialog.getSaveFileName(
            self, 'Export Wafer Map (Excel)', default_name, 'Excel Workbook (*.xlsx)'
        )
        if not out_path:
            return
        if not out_path.lower().endswith('.xlsx'):
            out_path += '.xlsx'

        mkey = self._current_mkey
        lo, hi, prod_lo, prod_hi = self._limits.get(mkey, (None, None, None, None))
        use_prod = self._use_prod_limits and (prod_lo is not None or prod_hi is not None)
        subs = self._all_subsites or [None]
        xs = [s['x'] for s in self._sites]
        ys = [s['y'] for s in self._sites]
        x0, x1 = min(xs), max(xs)
        y0, y1 = min(ys), max(ys)

        wb = Workbook()
        wb.remove(wb.active)
        thin = Side(style='thin', color='8899AA')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Excel palette constants (must match `_die_fill_hex()`).
        PASS = PASS_COLOR
        WARN = WARN_COLOR
        FAIL = FAIL_COLOR
        NEUTRAL = '#546e7a'
        NODATA = '#7b1fa2'

        def excel_text_color(hex_color: str) -> str:
            # Pick white text for darker fills; black text otherwise.
            h = hex_color.lstrip('#')
            r = int(h[0:2], 16)
            g = int(h[2:4], 16)
            b = int(h[4:6], 16)
            lum = (0.2126 * r + 0.7152 * g + 0.0722 * b) / 255.0
            return 'FFFFFF' if lum < 0.52 else '000000'

        for sub in subs:
            title = f'Design_{sub}' if sub is not None else 'Design_All'
            ws = wb.create_sheet(title=title[:31])
            ws['A1'] = 'Measurement'
            ws['B1'] = mkey
            ws['A2'] = 'Spec Limits'
            ws['B2'] = f'{lo if lo is not None else "—"} to {hi if hi is not None else "—"}'
            ws['A3'] = 'Prod Limits'
            ws['B3'] = (f'{prod_lo if prod_lo is not None else "—"} to {prod_hi if prod_hi is not None else "—"}'
                        if use_prod else 'off')
            ws['A4'] = 'Legend'
            ws['B4'] = 'Pass'
            ws['C4'] = 'Spec pass / Prod fail' if use_prod else 'Fail'
            ws['D4'] = 'Fail'
            ws['E4'] = 'No data'
            ws['B4'].fill = PatternFill(fill_type='solid', fgColor=PASS.replace('#', ''))
            ws['C4'].fill = PatternFill(fill_type='solid', fgColor=(WARN if use_prod else FAIL).replace('#', ''))
            ws['D4'].fill = PatternFill(fill_type='solid', fgColor=FAIL.replace('#', ''))
            ws['E4'].fill = PatternFill(fill_type='solid', fgColor=NODATA.replace('#', ''))

            ws['B4'].font = Font(bold=True, color=excel_text_color(PASS))
            ws['C4'].font = Font(bold=True, color=excel_text_color(WARN if use_prod else FAIL))
            ws['D4'].font = Font(bold=True, color=excel_text_color(FAIL))
            ws['E4'].font = Font(bold=True, color=excel_text_color(NODATA))
            for c in ('A1', 'A2', 'A3', 'A4'):
                ws[c].font = Font(bold=True)

            start_row = 7
            start_col = 3
            for x in range(x0, x1 + 1):
                cell = ws.cell(row=start_row - 1, column=start_col + (x - x0), value=str(x))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            for y in range(y0, y1 + 1):
                row = start_row + (y1 - y)
                cell = ws.cell(row=row, column=start_col - 1, value=str(y))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for s in self._sites:
                v = get_site_value(s, mkey, sub)
                row = start_row + (y1 - s['y'])
                col = start_col + (s['x'] - x0)
                cc = ws.cell(row=row, column=col, value=si_fmt(v) if v is not None else 'N/A')
                bg_hex = self._die_fill_hex(v, lo, hi, prod_lo, prod_hi, use_prod)
                hex_fill = bg_hex.replace('#', '')
                cc.fill = PatternFill(fill_type='solid', fgColor=hex_fill)
                cc.font = Font(color=excel_text_color(bg_hex))
                cc.border = border
                cc.alignment = Alignment(horizontal='center', vertical='center')

            for c in range(start_col, start_col + (x1 - x0 + 1)):
                ws.column_dimensions[get_column_letter(c)].width = 11
            for r in range(start_row, start_row + (y1 - y0 + 1)):
                ws.row_dimensions[r].height = 22

        try:
            wb.save(out_path)
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to save Excel file:\n{e}')
            return
        self.status.showMessage(f'  Exported wafer Excel map  ·  {out_path}')

    def _export_batch_report(self):
        if not self._batch_rows:
            QMessageBox.information(self, 'No batch data', 'Load a batch folder first.')
            return
        default_name = 'batch_report.xlsx'
        if self._batch_dir:
            default_name = f'{os.path.basename(self._batch_dir)}_batch_report.xlsx'
        out_path, _ = QFileDialog.getSaveFileName(
            self, 'Export Batch Report (Excel)', default_name, 'Excel Workbook (*.xlsx)'
        )
        if not out_path:
            return
        if not out_path.lower().endswith('.xlsx'):
            out_path += '.xlsx'

        mkey = self.batch_mkey_combo.currentText().strip()
        lo, hi, prod_lo, prod_hi = self._limits.get(mkey, (None, None, None, None))

        try:
            from openpyxl import Workbook
        except Exception:
            QMessageBox.warning(
                self,
                'Missing dependency',
                'Excel export requires openpyxl.\nInstall with: pip install openpyxl'
            )
            return

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        ws_summary.append(['Batch Folder', self._batch_dir or ''])
        ws_summary.append(['Measurement', mkey])
        ws_summary.append(['Spec Low', lo, 'Spec High', hi, 'Prod Low', prod_lo, 'Prod High', prod_hi])
        ws_summary.append([])
        ws_summary.append([
            'Wafer File', 'Lot', 'Designs', 'Design Used', 'Sites', 'Mean', 'Median', 'Std Dev',
            'Pass', 'Fail', 'Yield', 'Center Mean', 'Edge Mean', 'Edge-Center', 'Edge Fail %', 'Status'
        ])
        for r in self._batch_rows:
            edge_delta = (r['edge_mean'] - r['center_mean']) if (r['edge_mean'] is not None and r['center_mean'] is not None) else None
            ws_summary.append([
                r['rec']['name'], r['rec']['header'].get('LOT', '—'), r['designs'], r['design_used'], r['sites'],
                r['mean'], r['median'], r['std'], r['pass'], r['fail'], r['yield'],
                si_fmt(r['center_mean']) if r['center_mean'] is not None else 'N/A',
                si_fmt(r['edge_mean']) if r['edge_mean'] is not None else 'N/A',
                si_fmt(edge_delta) if edge_delta is not None else 'N/A',
                f'{r["edge_fail_pct"]:.1f}%' if r['edge_fail_pct'] is not None else 'N/A',
                r['status'],
            ])

        design_ids = sorted({sn for rec in self._batch_records for sn in rec.get('subs', [])})
        modes = [('All_Designs', None)] + [(f'Design_{sn}', sn) for sn in design_ids]
        headers = [
            'Wafer File', 'Lot', 'Design Used', 'Sites', 'Mean', 'Median', 'Std Dev',
            'Pass', 'Fail', 'Yield'
        ]
        for sheet_name, sub in modes:
            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append(['Batch Folder', self._batch_dir or ''])
            ws.append(['Measurement', mkey, 'Design Mode', 'All' if sub is None else str(sub)])
            ws.append([])
            ws.append(headers)
            for rec in self._batch_records:
                values = {s['name']: get_site_value(s, mkey, sub) for s in rec.get('sites', [])}
                vals = [v for v in values.values() if v is not None and math.isfinite(v)]
                mean_v = statistics.mean(vals) if vals else None
                med_v = statistics.median(vals) if vals else None
                std_v = statistics.pstdev(vals) if vals else None
                passed = 0
                for v in vals:
                    in_spec = (lo is None or v >= lo) and (hi is None or v <= hi)
                    in_prod = (prod_lo is None or v >= prod_lo) and (prod_hi is None or v <= prod_hi)
                    if in_spec and (in_prod if self._use_prod_limits else True):
                        passed += 1
                fail = max(0, len(vals) - passed)
                yld = (passed / len(vals) * 100.0) if vals else None
                ws.append([
                    rec['name'], rec['header'].get('LOT', '—'),
                    ('All' if sub is None else str(sub)), len(rec.get('sites', [])),
                    si_fmt(mean_v) if mean_v is not None else 'N/A',
                    si_fmt(med_v) if med_v is not None else 'N/A',
                    si_fmt(std_v) if std_v is not None else 'N/A',
                    passed, fail, (f'{yld:.1f}%' if yld is not None else 'N/A')
                ])

        try:
            wb.save(out_path)
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to write report:\n{e}')
            return
        self.status.showMessage(f'  Exported batch report  ·  {out_path}')

    # ── export ────────────────────────────────────────────────────────────────

    def export_map(self):
        if not self._sites:
            QMessageBox.information(self, 'Nothing to export',
                                    'Load a KDF file first.')
            return

        default_name = 'wafer_map.png'
        if self._filepath:
            base = os.path.splitext(os.path.basename(self._filepath))[0]
            default_name = f'{base}.png'

        path, selected_filter = QFileDialog.getSaveFileName(
            self, 'Export Wafer Map', default_name,
            'PNG Image (*.png);;JPEG Image (*.jpg)')
        if not path:
            return

        # Ensure a file extension so Qt can select an image writer reliably.
        lower = path.lower()
        if not (lower.endswith('.png') or lower.endswith('.jpg') or lower.endswith('.jpeg')):
            if 'jpeg' in (selected_filter or '').lower() or 'jpg' in (selected_filter or '').lower():
                path += '.jpg'
            else:
                path += '.png'

        SCALE = 3
        lw = max(1, self.canvas.width())
        lh = max(1, self.canvas.height())
        pw = lw * SCALE
        ph = lh * SCALE

        # Render directly to a high-resolution image for reliable export.
        img = QImage(pw, ph, QImage.Format_ARGB32)
        img.fill(Qt.transparent)
        painter = QPainter(img)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.TextAntialiasing)
        painter.scale(SCALE, SCALE)
        self.canvas.render(painter)
        painter.end()

        if path.lower().endswith(('.jpg', '.jpeg')):
            img = img.convertToFormat(QImage.Format_RGB32)

        fmt = b'PNG'
        if path.lower().endswith(('.jpg', '.jpeg')):
            fmt = b'JPEG'
        writer = QImageWriter(path, fmt)
        if fmt == b'JPEG':
            writer.setQuality(95)

        if writer.write(img):
            self.status.showMessage(
                f'  Exported {pw}×{ph} px  ·  {path}')
        else:
            QMessageBox.critical(
                self,
                'Export Error',
                f'Failed to save image.\n\n{writer.errorString()}'
            )

    # ── ui state ──────────────────────────────────────────────────────────────

    def _update_ui_state(self):
        has = bool(self._sites)
        self.design_combo.setEnabled(has)
        self.mkey_combo.setEnabled(has)
        self.continuous_heatmap_toggle.setEnabled(has)
        self.low_edit.setEnabled(has)
        self.high_edit.setEnabled(has)
        self.prod_toggle.setEnabled(has)
        self.prod_low_edit.setEnabled(has and self._use_prod_limits)
        self.prod_high_edit.setEnabled(has and self._use_prod_limits)

        # Keep batch UI interactive at all times; handlers validate data availability.
        self.batch_sections.setEnabled(True)
        self.batch_load_btn.setEnabled(True)
        self.batch_mkey_combo.setEnabled(True)
        self.batch_design_mode_combo.setEnabled(True)
        self.batch_sort_combo.setEnabled(True)
        self.batch_compare_btn.setEnabled(True)
        self.batch_export_report_btn.setEnabled(True)
        self.batch_golden_combo.setEnabled(True)
        self.batch_low_edit.setEnabled(True)
        self.batch_high_edit.setEnabled(True)
        self.batch_prod_toggle.setEnabled(True)
        self.batch_prod_low_edit.setEnabled(self._use_prod_limits)
        self.batch_prod_high_edit.setEnabled(self._use_prod_limits)
        self.batch_limits_apply_btn.setEnabled(True)
        self.batch_limits_clear_btn.setEnabled(True)

# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # Use a consistent DPI baseline so UI layout remains stable
    # across different display resolutions/scaling settings.
    QApplication.setAttribute(Qt.AA_Use96Dpi, True)
    app = QApplication(sys.argv)
    app.setApplicationName('Wafer Map Viewer')
    app.setStyle('Fusion')
    app.setWindowIcon(make_app_icon())
    win = MainWindow()
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        win._load_kdf(sys.argv[1])
    sys.exit(app.exec())

if __name__ == '__main__':
    main()

